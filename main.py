import re
import numpy as np
import pandas as pd
#from numba import jit
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from math import pi
from docx import Document
from docxtpl import DocxTemplate
from tkinter import *
from tkinter import filedialog as fd
from threading import *
import time
import snoop



КодKKS = 'KUR.0130.00USY.0.TZ.PA0025'
Архивный = ' '
дата = '01.11.2022'
Исполнитель = 'О.А. Тимофеева'
Наименование_работы = 'Эстакада технологических трубопроводов (00USY)\nЭстакада к зданию 20UMA (участок 1) \nСетевая вода от здания 20UMA к зданию 00UNA'
Исполнитель_должн = 'Разработал инж. 2к.'
КодKKSдокумента = 'KUR.0130.00USY.0.TK.TB0036'
пункт_графика = '2.28.2.1.8'
КодKKSспецификацииРД = 'KUR.0130.00USY.0.TK.TB0036.S0001'
КодKKSспецификацииарматурыРД = 'KUR.0130.00USY.0.TK.TB0036.S0002'
номерМДС = ''


def tkinter():
    def openfilepipe():
        global pipe_path
        pipe_path = fd.askopenfilename()
        LABEL14.config(text='Загружено')

    def openfilevalve():
        global valve_path
        valve_path = fd.askopenfilename()
        LABEL15.config(text='Загружено')

    def show():
        global пункт_графика
        global КодKKS
        global КодKKSдокумента
        global КодKKSспецификацииРД
        global КодKKSспецификацииарматурыРД
        global Архивный
        global номерМДС
        global Исполнитель_должн
        global Исполнитель
        global Наименование_работы
        global дата
        global вертуч1
        global вертуч2
        global вертуч3
        global вертуч4
        global вертуч5
        global вертуч6
        global вертуч7
        global вертуч8
        global вертуч9
        global вертуч10
        global вертуч11
        global вертуч12
        global вертуч13
        global вертуч14
        global вертуч15
        global вертуч16

        пункт_графика = entry1.get()
        КодKKS = entry2.get()
        КодKKSдокумента = entry3.get()
        КодKKSспецификацииРД = entry4.get()
        КодKKSспецификацииарматурыРД = entry5.get()
        Архивный = entry6.get()
        номерМДС = entry7.get()
        Исполнитель_должн = entry8.get()
        Исполнитель = entry9.get()
        Наименование_работы = entry10.get()
        дата = entry11.get()
        вертуч1 = entry12.get()
        вертуч2 = entry13.get()
        вертуч3 = entry14.get()
        вертуч4 = entry15.get()
        вертуч5 = entry16.get()
        вертуч6 = entry17.get()
        вертуч7 = entry18.get()
        вертуч8 = entry19.get()
        вертуч9 = entry20.get()
        вертуч10 = entry21.get()
        вертуч11 = entry22.get()
        вертуч12 = entry23.get()
        вертуч13 = entry24.get()
        вертуч14 = entry25.get()
        вертуч15 = entry26.get()
        вертуч16 = entry27.get()
        root.destroy()

    root = Tk()
    pipe_path = ''
    valve_path = ''
    entry1 = StringVar()
    entry2 = StringVar()
    entry3 = StringVar()
    entry4 = StringVar()
    entry5 = StringVar()
    entry6 = StringVar()
    entry7 = StringVar()
    entry8 = StringVar()
    entry9 = StringVar()
    entry10 = StringVar()
    entry11 = StringVar()
    entry12 = StringVar()
    entry13 = StringVar()
    entry14 = StringVar()
    entry15 = StringVar()
    entry16 = StringVar()
    entry17 = StringVar()
    entry18 = StringVar()
    entry19 = StringVar()
    entry20 = StringVar()
    entry21 = StringVar()
    entry22 = StringVar()
    entry23 = StringVar()
    entry24 = StringVar()
    entry25 = StringVar()
    entry26 = StringVar()
    entry27 = StringVar()
    root.title ("Расчет тепловой изоляции трубопроводов v1.0")
    root.geometry ("700x700")

    BTN_GO = Button (text = "Начать расчет",
                     padx="1",
                     pady="1",
                     font='Arial 13',
                     width="15",
                     command=show
                     )
    BTN_DWNLD1 = Button (text = "Загрузить", command=openfilepipe)
    BTN_DWNLD2 = Button (text = "Загрузить", command=openfilevalve)
    LABEL1 = Label(text="Введите пункт графика", justify=LEFT)
    LABEL2 = Label(text="Введите код KKS документа", justify=LEFT)
    LABEL3 = Label(text="Введите код KKS технологической части", justify=LEFT)
    LABEL4 = Label(text="Введите код KKS трубопроводной спецификации", justify=LEFT)
    LABEL5 = Label(text="Введите код КKS спецификации арматуры", justify=LEFT)
    LABEL6 = Label(text="Введите архивный номер", justify=LEFT)
    LABEL7 = Label(text="Введите номер сметы по МДС", justify=LEFT)
    LABEL8 = Label(text="Введите должность исполнителя", justify=LEFT)
    LABEL9 = Label(text="Введите ФИО исполнителя", justify=LEFT)
    LABEL10 = Label(text="Введите наименование работы", justify=LEFT)
    LABEL11 = Label(text="Введите дату", justify=LEFT)
    LABEL12 = Label(text="Загрузите файл спецификации комплекта РД на трубопроводы", wraplength="220", justify=LEFT)
    LABEL13 = Label(text="Загрузите файл спецификации комплекта РД на арматуру", wraplength="220", justify=LEFT)
    LABEL14 = Label(text="", wraplength="220", justify=LEFT)
    LABEL15 = Label(text="", wraplength="220", justify=LEFT)
    LABEL16 = Label(text="В окна ниже вставьте длины вертикальных участков, более 3 метров", wraplength="1000", justify=LEFT)
    ENTRY1 = Entry(root, textvariable=entry1, width="30")
    ENTRY2 = Entry(root, textvariable=entry2, width="30")
    ENTRY3 = Entry(root, textvariable=entry3, width="30")
    ENTRY4 = Entry(root, textvariable=entry4, width="30")
    ENTRY5 = Entry(root, textvariable=entry5, width="30")
    ENTRY6 = Entry(root, textvariable=entry6, width="30")
    ENTRY7 = Entry(root, textvariable=entry7, width="30")
    ENTRY8 = Entry(root, textvariable=entry8, width="30")
    ENTRY9 = Entry(root, textvariable=entry9, width="30")
    ENTRY10 = Entry(root, textvariable=entry10, width="30")
    ENTRY11 = Entry(root, textvariable=entry11, width="30")
    ENTRY12 = Entry(root, textvariable=entry12, width="10")
    ENTRY13 = Entry(root, textvariable=entry13, width="10")
    ENTRY14 = Entry(root, textvariable=entry14, width="10")
    ENTRY15 = Entry(root, textvariable=entry15, width="10")
    ENTRY16 = Entry(root, textvariable=entry16, width="10")
    ENTRY17 = Entry(root, textvariable=entry17, width="10")
    ENTRY18 = Entry(root, textvariable=entry18, width="10")
    ENTRY19 = Entry(root, textvariable=entry19, width="10")
    ENTRY20 = Entry(root, textvariable=entry20, width="10")
    ENTRY21 = Entry(root, textvariable=entry21, width="10")
    ENTRY22 = Entry(root, textvariable=entry22, width="10")
    ENTRY23 = Entry(root, textvariable=entry23, width="10")
    ENTRY24 = Entry(root, textvariable=entry24, width="10")
    ENTRY25 = Entry(root, textvariable=entry25, width="10")
    ENTRY26 = Entry(root, textvariable=entry26, width="10")
    ENTRY27 = Entry(root, textvariable=entry27, width="10")

    BTN_GO.place(relx=.36, rely=.9)
    BTN_DWNLD1.place (relx=.55, rely=.6)
    BTN_DWNLD2.place (relx=.55, rely=.65)
    LABEL1.place(relx=.1, rely=.05)
    LABEL2.place(relx=.1, rely=.1)
    LABEL3.place(relx=.1, rely=.15)
    LABEL4.place(relx=.1, rely=.2)
    LABEL5.place(relx=.1, rely=.25)
    LABEL6.place(relx=.1, rely=.3)
    LABEL7.place(relx=.1, rely=.35)
    LABEL8.place(relx=.1, rely=.4)
    LABEL9.place(relx=.1, rely=.45)
    LABEL10.place(relx=.1, rely=.5)
    LABEL11.place(relx=.1, rely=.55)
    LABEL12.place(relx=.1, rely=.6)
    LABEL13.place(relx=.1, rely=.65)
    LABEL14.place(relx=.7, rely=.6)
    LABEL15.place(relx=.7, rely=.65)
    LABEL16.place(relx=.1, rely=.75)

    ENTRY1.place(relx=.55, rely=.05)
    ENTRY2.place(relx=.55, rely=.1)
    ENTRY3.place(relx=.55, rely=.15)
    ENTRY4.place(relx=.55, rely=.2)
    ENTRY5.place(relx=.55, rely=.25)
    ENTRY6.place(relx=.55, rely=.3)
    ENTRY7.place(relx=.55, rely=.35)
    ENTRY8.place(relx=.55, rely=.4)
    ENTRY9.place(relx=.55, rely=.45)
    ENTRY10.place(relx=.55, rely=.5)
    ENTRY11.place(relx=.55, rely=.55)

    ENTRY12.place(relx=.1, rely=.8)
    ENTRY13.place(relx=.2, rely=.8)
    ENTRY14.place(relx=.3, rely=.8)
    ENTRY15.place(relx=.4, rely=.8)

    ENTRY16.place(relx=.5, rely=.8)
    ENTRY17.place(relx=.6, rely=.8)
    ENTRY18.place(relx=.7, rely=.8)
    ENTRY19.place(relx=.8, rely=.8)

    ENTRY20.place(relx=.1, rely=.85)
    ENTRY21.place(relx=.2, rely=.85)
    ENTRY22.place(relx=.3, rely=.85)
    ENTRY23.place(relx=.4, rely=.85)

    ENTRY24.place(relx=.5, rely=.85)
    ENTRY25.place(relx=.6, rely=.85)
    ENTRY26.place(relx=.7, rely=.85)
    ENTRY27.place(relx=.8, rely=.85)
    root.resizable(width=False, height=False)
    return root.mainloop()
##tkinter = tkinter()
@snoop
def find_prokladka(df):
    """
  Функция ищет все строчки в датафрейме, обозначающие начало группы.
  Далее, парсим каждую из строк и достаем тип прокладки и тип трубы
  и формируем словарь вида: индекс начала группы -> типы

  Пример:
    {
      0: ('надземная', 'прямой'),
      4: ('подземная', 'прямой'),
      9: ('надземная', 'обратной'),
      13: ('подземная', 'обратной')
    }
  """
    # Маска для поностью заполненных екселек
    # mask = df["Наименование и техническая   характеристика\n\n"].str.contains("прокладка")

    # Маска для файлика Ираклия
    mask = df["Примечание"].str.contains("Трубопровод")

    prokladka_dict = df[mask].to_dict("index")

    # print(prokladka_dict)

    prokladka_index_with_types = {}
    for index, value in prokladka_dict.items():

        # if "надземная" in value["Наименование и техническая   характеристика\n\n"]:
        #     ground_type = "надземная"
        # elif "подземная" in value["Наименование и техническая   характеристика\n\n"]:
        #     ground_type = "подземная"
        # else:
        #     raise ValueError(
        #         "В строчке" +
        #         value['Наименование и техническая   характеристика\n\n'] +
        #         "нет типа, проверьте данные"
        #     )

        if "подземная" in value["Наименование и техническая   характеристика\n\n"]:
            ground_type = "подземная"
        elif "Подземная" in value["Наименование и техническая   характеристика\n\n"]:
            ground_type = "подземная"
        else:  # Если тип не указан, то по умолчания считаем, что прокладка надземная
            ground_type = "надземная"

        if "прямой" in value["Примечание"]:
            pipeline_type = "прямой"
        elif "обратной" in value["Примечание"]:
            pipeline_type = "обратной"
        else:
            raise ValueError(
                "В строчке" +
                value['Примечание'] +
                "нет типа, проверьте данные"
            )

        temp = re.match(
            ".*t=(.*) °[С,C].*",
            value["Наименование и техническая   характеристика\n\n"]
        ).group(1)

        system = value['Примечание']

        prokladka_index_with_types[index] = (ground_type, pipeline_type, value["Код KKS"], temp, system)

    return prokladka_index_with_types
@snoop
def build_column_types(prokladka_index_with_types, df_length):
    index_list = list(prokladka_index_with_types.keys()) + [df_length]

    ground_type_column = []
    pipeline_column = []
    code_column = []
    system_name = []
    temperature = []

    for i in range(0, len(index_list) - 1):
        ground_type = prokladka_index_with_types[index_list[i]][0]
        pipeline_type = prokladka_index_with_types[index_list[i]][1]
        code_type = prokladka_index_with_types[index_list[i]][2]
        temp = prokladka_index_with_types[index_list[i]][3]
        system = prokladka_index_with_types[index_list[i]][4]

        ground_type_column.extend([ground_type] * (index_list[i + 1] - index_list[i]))
        pipeline_column.extend([pipeline_type] * (index_list[i + 1] - index_list[i]))
        code_column.extend([code_type] * (index_list[i + 1] - index_list[i]))
        temperature.extend([temp] * (index_list[i + 1] - index_list[i]))
        system_name.extend([system] * (index_list[i + 1] - index_list[i]))

    return ground_type_column, pipeline_column, code_column, temperature, system_name
@snoop
def pipelines_parse():
    # all_columns = pd.read_excel(
    #   "KUR.0130.00UNZ.SBA.TS.TB0045.S0001-MPA0001.xls",
    #   sheet_name="Sheet1",
    #   header=1
    # ).columns

    # print(f"Все колонки {all_columns}")

    my_df = (
        pd.read_excel(
            "input/KUR.0130.00USY.0.TK.TB0036.S0001-MPA0001.xls",
            sheet_name="Sheet1",
            header=1,
            usecols=['Код KKS', 'Наименование и техническая   характеристика\n\n', 'Примечание'],
        )
        .dropna(subset=['Наименование и техническая   характеристика\n\n', 'Примечание'])
        .reset_index(drop=True)
    )

    drenaz_index = list(my_df["Примечание"].values).index("Трубопроводы дренажей")
    # print(f"Индекс дренажа {drenaz_index}")
    my_df = (
        my_df
        .loc[:drenaz_index - 1]
        .reset_index(drop=True)
    )

    # Парсим типы
    types_dict = find_prokladka(my_df)
    # Собираем колонки с типами
    ground_type_column, pipeline_column, code_column, temperature, system = build_column_types(
        types_dict, len(my_df)
    )
    my_df.insert(0, "Тип прокладки", ground_type_column)
    my_df.insert(1, "Тип трубопровода", pipeline_column)
    my_df = my_df.drop("Код KKS", axis=1)
    my_df.insert(2, "Код KKS", code_column)
    my_df.insert(3, "Температура", temperature)
    my_df.insert(4, "Система", system)

    # Выкидываем лишнее
    my_df = my_df.loc[my_df["Примечание"].str.contains("строительная")]
    my_df = my_df.loc[~my_df["Примечание"].str.contains("Воздушник", case=False)]

    # print(my_df)
    # Парсим диаметр и длину
    pipe_parsed = my_df.copy()
    # print(my_df2.columns)
    pipe_parsed["Диаметр"] = (
        my_df["Наименование и техническая   характеристика\n\n"]
        .str.extract("([0-9]+)[х, x]").astype('int')
    )

    pipe_parsed["Длина"] = (
        my_df["Примечание"]
        .str.extract("([0-9]+[.,]?[0-9]*)")
    )
    pipe_parsed["Длина"] = (
        pipe_parsed["Длина"]
        .str.replace(",", ".")
        .astype('float')
    )

    pipe_parsed.index = ['{0}'.format(n) for n in range(pipe_parsed.index.shape[0])]
    pipe_parsed.rename(columns={'Наименование и техническая   характеристика\n\n': 'Наименование и техническая характеристика', 'Длина': 'Количество'}, inplace=True)
    pipe_parsed["Код KKS"] = pipe_parsed["Код KKS"].replace(r'\s+', '', regex=True)
    for i in range(len(pipe_parsed["Код KKS"])):
        pipe_parsed["Код KKS"][i] = pipe_parsed["Код KKS"][i].strip()
    ### Оставляем в наименоавании только "Труба" и диаметр
    for i in range(len(pipe_parsed['Наименование и техническая характеристика'])):
        end = pipe_parsed['Наименование и техническая характеристика'][i].index("СТО")
        pipe_parsed['Наименование и техническая характеристика'][i] = pipe_parsed['Наименование и техническая характеристика'][i][:end]
    pipe_parsed.to_excel("pipe_parsed.xlsx")
    return pipe_parsed

#pipe_parsed = pipelines_parse()
@snoop
def parse_note(row):
    if "подземная" in row["Примечание"].lower():
        return "подземная"

    return "надземная"
# all_columns = pd.read_excel(
#   "KUR.0130.00UNZ.SBA.TS.TB0051.S0002-MPA0001.xls",
#   sheet_name="CommonList",
#   header=1
# ).columns
@snoop
def armatura_presence():
    print('Есть ли арматура в комплекте? Напишите "да", если есть', sep='\n')
    armatura_presence = str(input())
    return armatura_presence
@snoop
def parse_armatura():
    if armatura_presence != "да":
        return None
    else:
        armatura = (
            pd.read_excel(
                "input/KUR.0130.00USY.0.TK.TB0036.S0002-MPA0001.xlsx",
                sheet_name="CommonList",
                header=1,
                usecols=['Код KKS', 'Наименование и техническая   характеристика', 'Примечание']
            )
            .dropna()
            .reset_index(drop=True)
        )

        armatura["Диаметр"] = (
            armatura["Наименование и техническая   характеристика"]
            .str.extract("DN\s([0-9]+);").astype('int')
        )
        armatura["Наименование и техническая   характеристика"] = (
            armatura["Наименование и техническая   характеристика"]
            .str.extract('([^0-9;]+)').astype('str')

        )


        armatura["Тип прокладки"] = armatura.apply(parse_note, axis=1)

        # Парсим  код арматуры
        armatura.insert(1, "Постфикс Код KKS", armatura["Код KKS"].str[7:])
        armatura["Код KKS"] = armatura["Код KKS"].str[:7]

        # Фильтруеми воздушники
        #armatura = armatura.loc[~armatura["Постфикс Код КодKKSдокумента"].str.contains("AA5")]
        armatura["Количество"] = 1

        armatura = armatura.groupby(by=["Код KKS", "Диаметр", "Тип прокладки", "Наименование и техническая   характеристика",],
                                    as_index=False)["Количество"].sum()
        armatura.rename(
            columns={'Наименование и техническая   характеристика': 'Наименование и техническая характеристика'},
            inplace=True)

        #Для того, чтобы получить корректное значение температруы для арматуры - не парсим это значение из спецификации,
        # а передаем "в наследство" от трубы. Чтобы реализовать это - создаем маленький датафрейм, где всего
        # 2 столбца - Код KKS, температура (повторы выкинуты)- это срез из распарсенного экселя труб. И по соотвествию
        # системы проставляем  #температуру для арматуры.
        armatura["Температура"] = np.nan
        armatura["Система"] = np.nan
        pipe_parsed_slice = pd.DataFrame(data=pipe_parsed['Код KKS'])
        pipe_parsed_slice.insert(1, "Температура", pipe_parsed['Температура'])
        pipe_parsed_slice.insert(2, "Система", pipe_parsed['Система'])
        pipe_parsed_slice = pipe_parsed_slice.drop_duplicates()
        #pipe_parsed_slice.to_excel("pipe_parsed_slice.xlsx", index=False)
        armatura.to_excel("armatura_parsed.xlsx", index=False)
        j = 0
        print(pipe_parsed_slice["Код KKS"][0])
        for i in range(len(armatura["Код KKS"])):
            if armatura["Код KKS"][i] == pipe_parsed_slice["Код KKS"][j]:
                armatura["Температура"][i] = pipe_parsed_slice["Температура"][j]
                armatura["Система"][i] = pipe_parsed_slice["Система"][j]
            else:
                armatura["Температура"][i] = pipe_parsed_slice["Температура"][j+1]
                armatura["Система"][i] = pipe_parsed_slice["Система"][j+1]
                j += 1

    #armatura.to_excel("armatura_parsed.xlsx", index=False)
        return armatura

#armatura_parsed = parse_armatura()
### Создаем таблицу расходов материалов для расчета арматуры DN20 - DN800
### Расчет позиций 1-5 для арматуры DN2-DN800
@snoop
def armatura_pos1_5():
    if armatura_presence != "да":
        return None
    else:
        armatura_pos1_5_data = {
            'Диаметр': [15, 15, 20, 20, 25, 25, 32, 32, 40, 40, 50, 50, 65, 65, 80, 80, 100, 100, 125, 125, 150, 150, 200, 200,
            250, 250, 300, 300, 350, 350, 400, 400, 500, 500, 600, 600, 700, 700, 800, 800],
            'δк': [40, 60, 40, 60, 40, 60, 40, 60, 40, 60, 40, 60, 40, 60, 40, 60, 40, 60, 40, 60, 40, 60, 40, 60, 40, 60, 40,
            60, 40, 60, 40, 60, 40, 60, 40, 60, 40, 60, 40, 60],
            'Тип прокладки': ['подземная', 'надземная', 'подземная', 'надземная', 'подземная', 'надземная', 'подземная',
            'надземная', 'подземная', 'надземная', 'подземная', 'надземная', 'подземная', 'надземная', 'подземная', 'надземная',
            'подземная', 'надземная', 'подземная', 'надземная', 'подземная', 'надземная', 'подземная', 'надземная', 'подземная',
            'надземная', 'подземная', 'надземная', 'подземная', 'надземная', 'подземная', 'надземная', 'подземная', 'надземная',
            'подземная', 'надземная', 'подземная', 'надземная','подземная', 'надземная'],
            #'Поверхность, м2, ПОЗ.1':[0.158, 0.185, 0.18, 0.208, 0.199, 0.23, 0.26, 0.3, 0.29, 0.33, 0.3, 0.34, 0.34, 0.38,
             #0.41, 0.45, 0.47, 0.52, 0.61, 0.66],
            #'Поверхность, м2, ПОЗ.2': [0.018, 0.027, 0.02, 0.03, 0.021, 0.031, 0.02, 0.03, 0.03, 0.04, 0.03, 0.04, 0.03, 0.04,
            #0.03, 0.05, 0.04, 0.05, 0.04, 0.06],
            #'Поверхность, м2, ПОЗ.3': [0.004, 0.005, 0.004, 0.005, 0.004, 0.005, 0.004, 0.005, 0.004, 0.005, 0.004, 0.005, 0.004,
            #0.005, 0.004, 0.005, 0.004, 0.005, 0.004, 0.005],
            #'Объем, м3, ПОЗ.5': [0.005, 0.009, 0.005, 0.01, 0.007, 0.012, 0.014, 0.026, 0.017, 0.028, 0.018, 0.03, 0.02, 0.033,
            # 0.024, 0.04, 0.028, 0.047, 0.037, 0.06],  #
            'V, м3': [0.01, 0.018, 0.01, 0.018, 0.01, 0.018, 0.01, 0.02, 0.014, 0.024, 0.028, 0.052, 0.034, 0.056, 0.036, 0.06,
            0.04, 0.066, 0.048, 0.08, 0.056, 0.094, 0.074, 0.12, 0.046, 0.072, 0.054, 0.084, 0.064, 0.1, 0.076, 0.12, 0.1, 0.15,
            0.13, 0.2, 0.19, 0.32, 0.21, 0.32],  ##########
            'S, м2': [0.420, 0.518, 0.420, 0.518, 0.420, 0.518, 0.472, 0.576, 0.514, 0.624, 0.632, 0.760, 0.732, 0.860, 0.752,
            0.880, 0.832, 0.960, 0.972, 1.140, 1.132, 1.280, 1.412, 1.600, 1.556, 1.76, 1.816, 2.008, 2.100, 2.324, 2.404, 2.636,
            3.120, 3.384, 3.916, 4.196, 5.07, 6.11, 5.64, 6.73]
        }
        armatura_pos1_5 = pd.DataFrame(armatura_pos1_5_data)
        return armatura_pos1_5

#armatura_pos1_5 = armatura_pos1_5()
@snoop
def armatura_fasteners():
    if armatura_presence != "да":
        return None
    else:
        armatura_fasteners_data = {
            'Диаметр':[15, 15, 20, 20, 25, 25, 32, 32, 40, 40, 50, 50, 65, 65, 80, 80, 100, 100, 125, 125, 150, 150, 200, 200,
            250, 250, 300, 300, 350, 350, 400, 400, 500, 500, 600, 600, 700, 700, 800, 800],
            'Тип прокладки': ['подземная', 'надземная','подземная', 'надземная','подземная', 'надземная', 'подземная', 'надземная',
            'подземная', 'надземная', 'подземная', 'надземная', 'подземная', 'надземная', 'подземная', 'надземная', 'подземная',
            'надземная', 'подземная', 'надземная', 'подземная', 'надземная', 'подземная', 'надземная', 'подземная', 'надземная',
            'подземная', 'надземная', 'подземная', 'надземная', 'подземная', 'надземная', 'подземная', 'надземная', 'подземная',
            'надземная', 'подземная', 'надземная', 'подземная', 'надземная'],
            'δк': [40, 60, 40, 60, 40, 60, 40, 60, 40, 60, 40, 60, 40, 60, 40, 60, 40, 60, 40, 60, 40, 60, 40, 60, 40, 60, 40,
            60, 40, 60, 40, 60, 40, 60, 40, 60, 40, 60, 40, 60],
            'Лента АД1.08х20': [0.066, 0.076, 0.066, 0.076, 0.066, 0.076, 0.071, 0.082, 0.073, 0.084, 0.078, 0.097, 0.083, 0.106,
            0.097, 0.106, 0.097, 0.115, 0.106, 0.123, 0.115, 0.132, 0.132, 0.141, 0,0,0,0,0,0,0,0,0,0,0,0,0.37,0.39,0.37,0.39],
            'Проволока 0,8-0-Ч': [0.01, 0.02, 0.01, 0.02, 0.01, 0.02, 0.02, 0.024, 0.02, 0.022, 0.02, 0.026, 0.024, 0.028, 0.024,
            0.03, 0.026, 0.032, 0.032, 0.038, 0.038, 0.042, 0.046, 0.052, 0.03, 0.034, 0.037, 0.04, 0.044, 0.046, 0.05, 0.053,
            0.066, 0.069, 0.084, 0.087,0.11,0.11,0.11,0.11],
            'Пряжка тип II-А': [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 0, 0, 0, 0, 0, 0, 0, 0,
            0, 0, 0, 0, 0, 0, 0, 0],
            'Заклепка комбинированная ЗК-12-4,5':[8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 12, 8, 12, 12, 12, 12, 12,
            34, 40, 40, 40, 40, 40, 46, 46, 46, 52, 52, 52,60,60,60,60],
            'Крючок':[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0.056,0.056,0.056,0.056,0.056,0.056,0.056,
            0.056,0.056,0.056,0.056,0.056,0.07,0.07,0.07,0.07],
            'Серьга':[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0.044,0.044,0.044,0.044,0.044,0.044,0.044,0.044,
            0.044,0.044,0.044,0.044,0.055,0.055,0.055,0.055],
            'Рычаг':[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0.12,0.12,0.12,0.12,0.12,0.12,0.12,0.12,0.12,
            0.12,0.12,0.12,0.15,0.15,0.15,0.15],
            'Основание':[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0.064,0.064,0.064,0.064,0.064,0.064,0.064,
            0.064,0.064,0.064,0.064,0.064,0.08,0.08,0.08,0.08],
            'Заклепка 4x24.37':[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0.056,0.056,0.056,0.056,0.056,0.056,0.056,
            0.056,0.056,0.056,0.056,0.056,0.07,0.07,0.07,0.07]
        }
        armatura_fasteners = pd.DataFrame(armatura_fasteners_data)
        return armatura_fasteners

#armatura_fasteners = armatura_fasteners()
# не забыть пересчитать Лак БТ-7 в м2, формула такая: массу ленты 0,7х20/5,495, лак используется только для подземной прокладки(видимо красим ленту)
### расчет изоляции
@snoop
def pipeline_fasteners():
    pipeline_fasteners_data = {
        'Диаметр': [57, 76, 89, 108, 133, 133, 159, 159, 219, 219, 219, 273, 273, 273, 325, 325, 325, 373, 373, 373, 426,
        426, 426, 480, 480, 480, 530, 530, 530, 630, 630, 630, 720, 720, 720, 820, 820, 820],
        'δк': [50, 50, 50, 50, 50, 60, 50, 60, 50, 60, 70, 50, 60, 70, 50, 60, 70, 50, 60, 70, 50, 60, 70, 60, 70, 80, 60,
        70, 80, 60, 70, 80, 60, 70, 80, 60, 70, 80],
        'Лента 0,7х20': [0.210, 0.225, 0.240, 0.260, 0.285, 0.310, 0.310, 0.330, 0.370, 0.390, 0.415, 0.430, 0.450, 0.480,
        0.490, 0.510, 0.530, 0.540, 0.560, 0.580, 0.590, 0.610, 0.630, 0.660, 0.680, 0.700, 0.720, 0.740, 0.760, 0.82, 0.84,
        0.86, 0.92, 0.94, 0.96, 1.02, 1.04, 1.06],
        'Винт самонарезной': [7, 7, 7, 7, 7, 7, 7, 7, 7, 7, 11, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 18.2, 18.2,
        18.2, 18.2, 22.05, 25.9, 25.9, 25.9, 25.9, 26.6, 26.6, 26.6, 27.3, 27.3, 27.3],
        'Проволока 2-0-Ч': [0.022, 0.0255, 0.0275, 0.0305, 0.035, 0.0360, 0.0390, 0.040, 0.0450, 0.0460, 0.0805, 0.0565,
        0.0580, 0.0965, 0.0655, 0.0670, 0.1095, 0.0740, 0.0760, 0.1225, 0.0830, 0.1335, 0.1840, 0.0920, 0.1460, 0.020, 0.010,
        0.1585, 0.2170, 0.116, 0.182, 0.248, 0.131, 0.204, 0.277, 0.147, 0.2275, 0.308],
        'Проволока 0,8-0-Ч': [0.0075, 0.008, 0.0095, 0.011, 0.011, 0.012, 0.013, 0.014, 0.0175, 0.018, 0.0185, 0.0285, 0.03,
        0.0315, 0.0325, 0.034, 0.0355, 0.0355, 0.037, 0.039, 0.0395, 0.041, 0.0425, 0.045, 0.0465, 0.048, 0.051, 0.0525,
        0.054, 0.053, 0.0545, 0.056, 0.62, 0.635, 0.65, 0.69, 0.705, 0.72],
        'Скоба опорная': [4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 3, 3, 3, 3, 3, 3, 4,
        4, 4, 4, 4, 4],
        'Пряжка тип I-О': [3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3,
        3, 3, 3, 3, 3, 3],
        'Лак БТ-7,кг': [0.006, 0.007, 0.007, 0.008, 0.006, 0.008, 0.009, 0.01, 0.011, 0.011, 0.012, 0.013, 0.013, 0.013,
        0.014, 0.015, 0.015, 0.016, 0.016, 0.016, 0.017, 0.018, 0.018, 0.019, 0.020, 0.021, 0.021, 0.0215, 0.0220, 0.024,
        0.0245, 0.025, 0.026, 0.027, 0.028, 0.030, 0.0305, 0.031],
        'Лента 2х30': [0.152, 0.194, 0.199, 0.227, 0.264, 0.272, 0.302, 0.310, 0.0595, 0.074, 0.087, 0.0595, 0.074, 0.087,
        0.0595, 0.074, 0.087, 0.0595, 0.074, 0.087, 0.100, 0.120, 0.140, 0.100, 0.120, 0.140, 0.120, 0.140, 0.160, 0.120,
        0.140, 0.160, 0.120, 0.140, 0.160, 0.120, 0.140, 0.160],
        'Лента 3х30': [0, 0, 0, 0, 0, 0, 0, 0, 0.223, 0.223, 0.223, 0.284, 0.284, 0.284, 0.341, 0.341, 0.341, 0.399, 0.399,
        0.399, 0.453, 0.453, 0.453, 0.513, 0.513, 0.513, 0.568, 0.568, 0.568, 0.679, 0.679, 0.679, 0.779, 0.779, 0.779, 0.890,
        0.890, 0.890],
        'Уголок 30х30x3': [0, 0, 0, 0, 0, 0, 0, 0, 0.114, 0.114, 0.114, 0.114, 0.114, 0.114, 0.114, 0.114, 0.114, 0.114, 0.114,
        0.114, 0.114, 0.114, 0.114, 0.114, 0.114, 0.114, 0.114, 0.114, 0.114, 0.114, 0.114, 0.114, 0.114, 0.114, 0.114, 0.114,
        0.114, 0.114],
        'Скоба навесная': [20, 20, 20, 20, 20, 20, 20, 20, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 40, 40,
        40, 40, 40, 40, 40, 40, 40, 50, 50, 50, 50, 50, 60],
        'Лист АД1.Н-0.8': [0.04, 0.07, 0.07, 0.08, 0.095, 0.12, 0.105, 0.13, 0.13, 0.16, 0.195, 0.165, 0.2, 0.24, 0.19, 0.23,
        0.275, 0.21, 0.26, 0.305, 0.23, 0.28, 0.335, 0.32, 0.375, 0.43, 0.34, 0.405 ,0.47, 0.4, 0.475, 0.55, 0.61, 0.7, 0.79,
        0.6, 0.65, 0.7],
        'Болт М8': [0.051, 0.051, 0.051, 0.051, 0.051, 0.051, 0.051, 0.051, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
        0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        'Болт М12': [0, 0, 0, 0, 0, 0, 0, 0, 0.372, 0.372, 0.372, 0.372, 0.372, 0.372, 0.372, 0.372, 0.372, 0.372, 0.372, 0.372,
        0.372, 0.372, 0.372, 0.372, 0.372, 0.372, 0.372, 0.372, 0.372, 0.372, 0.372, 0.372, 0.372, 0.372, 0.372, 0.372, 0.372,
        0.372],
        'Гайка М8': [0.015, 0.015, 0.015, 0.015, 0.015, 0.015, 0.015, 0.015, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
        0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        'Гайка М12': [0, 0, 0, 0, 0, 0, 0, 0, 0.090, 0.090, 0.090, 0.090, 0.090, 0.090, 0.090, 0.090, 0.090, 0.090, 0.090, 0.090,
        0.090, 0.090, 0.090, 0.090, 0.090, 0.090, 0.090, 0.090, 0.090, 0.090, 0.090, 0.090, 0.090, 0.090, 0.090, 0.090, 0.090,
        0.090]
    }
    pipeline_fasteners = pd.DataFrame(pipeline_fasteners_data)
    return pipeline_fasteners
#pipeline_fasteners = pipeline_fasteners()
### Готовим список длин вертикальных участков трассы больше 3 метров.
@snoop
def ascents():
    list_of_types = []
    column_list = list(pipe_parsed.columns)
    slise_start = column_list.index('Тип прокладки')
    slise_finish = column_list.index('Температура')
    pipe_parsed_slise = pd.DataFrame(data = pipe_parsed.iloc[:,slise_start:slise_finish])
    pipe_parsed_slise.insert(3, "Диаметр", pipe_parsed['Диаметр'])
    list_of_types = pipe_parsed_slise.values.tolist()
    n = len(list_of_types)
    ascent_list = []

    #Мой старый вариант
    print('Есть ли вертикальные участки трубопроводов больше 3 метров длиной? Напишите "да", если есть', sep = '\n')
    answer = str(input())
    # answer = 0
    if answer == 'да':
        print(list_of_types)
        print(len(list_of_types))
        for i in range(0, n):
            dlina_asc_part = float(input())
            ascent_list.append(dlina_asc_part)
    else:
        for i in range(0, n):
            dlina_asc_part = 0
            ascent_list.append(dlina_asc_part)

    return ascent_list
    ## Какая-то фигня от Андрея

    # ascent_list_str = [вертуч1, вертуч2, вертуч3, вертуч4, вертуч5, вертуч6, вертуч7, вертуч8, вертуч9, вертуч10,
    #                    вертуч11, вертуч12, вертуч13, вертуч14, вертуч15, вертуч16]
    # for i in range(16):
    #     dlina_asc_part = float(ascent_list_str[i])
    #     ascent_list.append(dlina_asc_part)

    #return ascent_list

#ascent_list = ascents()

### В цикле пробегаем по распарсенному экселю, определяем толщину изоляции в конструкции, δк
### Расчет и определение материалов ДЛЯ ИЗГОТОВЛЕНИЯ изоляции тут нам поможет формула определения расхода материала
### для изготовления матов и покровного слоя исходя из диаметра и толщины изоляции. Материал (Т-23) для матов
### в облкадках/ без облкадок так же определяем здесь

@snoop

def pipeline_insulation_thickness():
    thickness_list = []
    pipe_parsed["steel_0_5mm"] = np.nan
    pipe_parsed["steel_0_8mm"] = np.nan
    pipe_parsed["mati_2m100_80mm"] = np.nan
    pipe_parsed["mati_2m100_70mm"] = np.nan
    pipe_parsed["mati_2m100_60mm"] = np.nan
    pipe_parsed["mati_2m100_50mm"] = np.nan
    pipe_parsed["mati_m100_70mm"] = np.nan
    pipe_parsed["mati_m100_60mm"] = np.nan
    pipe_parsed["mati_m100_50mm"] = np.nan
    pipe_parsed["protective_cover_T_23"] = np.nan
    for i in range(len(pipe_parsed)):
        if pipe_parsed['Тип прокладки'][i] == 'надземная' and pipe_parsed['Тип трубопровода'][i] == 'прямой':
            if pipe_parsed['Диаметр'][i] <= 108:
                thickness = 50
                thickness_list.append(thickness)
                mati_m100_50 = round((thickness / 1000) * pi * (pipe_parsed['Диаметр'][i] / 1000 + thickness / 1000),3)
                pipe_parsed["mati_m100_50mm"][i] = mati_m100_50
                steel = round(pi * (pipe_parsed['Диаметр'][i] / 1000 + (thickness / 1000) * 2), 3)
                pipe_parsed["steel_0_5mm"][i] = steel
            elif 108 < pipe_parsed['Диаметр'][i] <= 159:
                thickness = 60
                thickness_list.append(thickness)
                mati_m100_60 = round((thickness / 1000) * pi * (pipe_parsed['Диаметр'][i] / 1000 + thickness / 1000),3)
                pipe_parsed["mati_m100_60mm"][i] = mati_m100_60
                steel = round(pi * (pipe_parsed['Диаметр'][i] / 1000 + (thickness / 1000) * 2), 3)
                pipe_parsed["steel_0_5mm"][i] = steel
            elif 159 < pipe_parsed['Диаметр'][i] < 250:
                thickness = 70
                thickness_list.append(thickness)
                mati_m100_70 = round((thickness / 1000) * pi * (pipe_parsed['Диаметр'][i] / 1000 + thickness / 1000),3)
                pipe_parsed["mati_m100_70mm"][i] = mati_m100_70
                steel = round(pi * (pipe_parsed['Диаметр'][i] / 1000 + (thickness / 1000) * 2), 3)
                pipe_parsed["steel_0_5mm"][i] = steel
            elif 250 <= pipe_parsed['Диаметр'][i] < 480:
                thickness = 70
                thickness_list.append(thickness)
                mati_2m100_70 = round((thickness / 1000) * pi * (pipe_parsed['Диаметр'][i] / 1000 + thickness / 1000),3)
                pipe_parsed["mati_2m100_70mm"][i] = mati_2m100_70
                steel = round(pi * (pipe_parsed['Диаметр'][i] / 1000 + (thickness / 1000) * 2), 3)
                pipe_parsed["steel_0_5mm"][i] = steel
            elif 480 <= pipe_parsed['Диаметр'][i] < 600:
                thickness = 80
                thickness_list.append(thickness)
                mati_2m100_80 = round((thickness / 1000) * pi * (pipe_parsed['Диаметр'][i] / 1000 + thickness / 1000), 3)
                pipe_parsed["mati_2m100_80mm"][i] = mati_2m100_80
                steel = round(pi * (pipe_parsed['Диаметр'][i] / 1000 + (thickness / 1000) * 2), 3)
                pipe_parsed["steel_0_5mm"][i] = steel
            elif 600 <= pipe_parsed['Диаметр'][i]:
                thickness = 80
                thickness_list.append(thickness)
                mati_2m100_80 = round((thickness / 1000) * pi * (pipe_parsed['Диаметр'][i] / 1000 + thickness / 1000), 3)
                pipe_parsed["mati_2m100_80mm"][i] = mati_2m100_80
                steel = round(pi * (pipe_parsed['Диаметр'][i] / 1000 + (thickness / 1000) * 2), 3)
                pipe_parsed["steel_0_8mm"][i] = steel
        elif pipe_parsed['Тип прокладки'][i] == 'надземная' and pipe_parsed['Тип трубопровода'][i] == 'обратной':
            if pipe_parsed['Диаметр'][i] < 250:
                thickness = 50
                thickness_list.append(thickness)
                mati_m100_50 = round((thickness / 1000) * pi * (pipe_parsed['Диаметр'][i] / 1000 + thickness / 1000), 3)
                pipe_parsed["mati_m100_50mm"][i] = mati_m100_50
                steel = round(pi * (pipe_parsed['Диаметр'][i] / 1000 + (thickness / 1000) * 2), 3)
                pipe_parsed["steel_0_5mm"][i] = steel
            elif  250 <= pipe_parsed['Диаметр'][i] <= 426:
                thickness = 50
                thickness_list.append(thickness)
                mati_2m100_50 = round((thickness / 1000) * pi * (pipe_parsed['Диаметр'][i] / 1000 + thickness / 1000), 3)
                pipe_parsed["mati_2m100_50mm"][i] = mati_2m100_50
                steel = round(pi * (pipe_parsed['Диаметр'][i] / 1000 + (thickness / 1000) * 2), 3)
                pipe_parsed["steel_0_5mm"][i] = steel
            elif 480 <= pipe_parsed['Диаметр'][i] < 600:
                thickness = 60
                thickness_list.append(thickness)
                mati_2m100_60 = round((thickness / 1000) * pi * (pipe_parsed['Диаметр'][i] / 1000 + thickness / 1000), 3)
                pipe_parsed["mati_2m100_60mm"][i] = mati_2m100_60
                steel = round(pi * (pipe_parsed['Диаметр'][i] / 1000 + (thickness / 1000) * 2), 3)
                pipe_parsed["steel_0_5mm"][i] = steel
            elif 600 <= pipe_parsed['Диаметр'][i]:
                thickness = 60
                thickness_list.append(thickness)
                mati_2m100_60 = round((thickness / 1000) * pi * (pipe_parsed['Диаметр'][i] / 1000 + thickness / 1000), 3)
                pipe_parsed["mati_2m100_60mm"][i] = mati_2m100_60
                steel = round(pi * (pipe_parsed['Диаметр'][i] / 1000 + (thickness / 1000) * 2), 3)
                pipe_parsed["steel_0_8mm"][i] = steel
    ###### расчет для подземной прокладки
        elif pipe_parsed['Тип прокладки'][i] == 'подземная':
            if pipe_parsed['Диаметр'][i] <= 133:
                thickness = 50
                thickness_list.append(thickness)
                mati_m100_50 = round((thickness / 1000) * pi * (pipe_parsed['Диаметр'][i] / 1000 + thickness / 1000), 3)
                pipe_parsed["mati_m100_50mm"][i] = mati_m100_50
                S = round(pi * (pipe_parsed['Диаметр'][i] / 1000 + (thickness / 1000) * 2), 3)
                pipe_parsed["protective_cover_T_23"][i] = S
            elif 133 < pipe_parsed['Диаметр'][i] < 250:
                thickness = 60
                thickness_list.append(thickness)
                mati_m100_60 = round((thickness / 1000) * pi * (pipe_parsed['Диаметр'][i] / 1000 + thickness / 1000), 3)
                pipe_parsed["mati_m100_60mm"][i] = mati_m100_60
                S = round(pi * (pipe_parsed['Диаметр'][i] / 1000 + (thickness / 1000) * 2), 3)
                pipe_parsed["protective_cover_T_23"][i] = S
            elif 250 <= pipe_parsed['Диаметр'][i] <= 426:
                thickness = 60
                thickness_list.append(thickness)
                mati_2m100_60 = round((thickness / 1000) * pi * (pipe_parsed['Диаметр'][i] / 1000 + thickness / 1000), 3)
                pipe_parsed["mati_2m100_60mm"][i] = mati_2m100_60
                S = round(pi * (pipe_parsed['Диаметр'][i] / 1000 + (thickness / 1000) * 2), 3)
                pipe_parsed["protective_cover_T_23"][i] = S
            elif 426 < pipe_parsed['Диаметр'][i]:
               thickness = 70
               thickness_list.append(thickness)
               mati_2m100_70 = round((thickness / 1000) * pi * (pipe_parsed['Диаметр'][i] / 1000 + thickness / 1000), 3)
               pipe_parsed["mati_2m100_70mm"][i] = mati_2m100_70
               S = round(pi * (pipe_parsed['Диаметр'][i] / 1000 + (thickness / 1000) * 2), 3)
               pipe_parsed["protective_cover_T_23"][i] = S

    pipe_parsed.insert(7, "δк", thickness_list)
    #print(thickness_list)
    return pipe_parsed

#pipe_parsed = pipeline_insulation_thickness()
@snoop
def pipeline_insun_preporation():

    #pipe_parsed.insert(7, "δк", thickness_list)
    pipe_parsed.insert(8, "Длина вертикальных участков", ascent_list)
    ### Умножаем расходы для матов и для покровного слоя на длину трубопроводов и заносим в таблицу
    pipe_parsed["steel_0_5mm * Количество"] = pipe_parsed["steel_0_5mm"] * pipe_parsed["Количество"]
    pipe_parsed["steel_0_5mm * Количество"] = pipe_parsed["steel_0_5mm * Количество"].round(3)
    pipe_parsed["steel_0_8mm * Количество"] = pipe_parsed["steel_0_8mm"] * pipe_parsed["Количество"]
    pipe_parsed["steel_0_8mm * Количество"] = pipe_parsed["steel_0_8mm * Количество"].round(3)
    pipe_parsed['mati_2m100_80mm * Количество'] = pipe_parsed['Количество'] * pipe_parsed['mati_2m100_80mm']
    pipe_parsed['mati_2m100_80mm * Количество'] = pipe_parsed['mati_2m100_80mm * Количество'].round(3)
    pipe_parsed['mati_2m100_70mm * Количество'] = pipe_parsed['Количество'] * pipe_parsed['mati_2m100_70mm']
    pipe_parsed['mati_2m100_70mm * Количество'] = pipe_parsed['mati_2m100_70mm * Количество'].round(3)
    pipe_parsed['mati_2m100_60mm * Количество'] = pipe_parsed['Количество'] * pipe_parsed['mati_2m100_60mm']
    pipe_parsed['mati_2m100_60mm * Количество'] = pipe_parsed['mati_2m100_60mm * Количество'].round(3)
    pipe_parsed['mati_2m100_50mm * Количество'] = pipe_parsed['Количество'] * pipe_parsed['mati_2m100_50mm']
    pipe_parsed['mati_2m100_50mm * Количество'] = pipe_parsed['mati_2m100_50mm * Количество'].round(3)
    pipe_parsed['mati_m100_70mm * Количество'] = pipe_parsed['Количество'] * pipe_parsed['mati_m100_70mm']
    pipe_parsed['mati_m100_70mm * Количество'] = pipe_parsed['mati_m100_70mm * Количество'].round(3)
    pipe_parsed['mati_m100_60mm * Количество'] = pipe_parsed['Количество'] * pipe_parsed['mati_m100_60mm']
    pipe_parsed['mati_m100_60mm * Количество'] = pipe_parsed['mati_m100_60mm * Количество'].round(3)
    pipe_parsed['mati_m100_50mm * Количество'] = pipe_parsed['Количество'] * pipe_parsed['mati_m100_50mm']
    pipe_parsed['mati_m100_50mm * Количество'] = pipe_parsed['mati_m100_50mm * Количество'].round(3)
    pipe_parsed['T_23_2m100_80mm'] = pipe_parsed['mati_2m100_80mm * Количество'] * 27
    pipe_parsed['T_23_2m100_70mm'] = pipe_parsed['mati_2m100_70mm * Количество'] * 30
    pipe_parsed['T_23_2m100_60mm'] = pipe_parsed['mati_2m100_60mm * Количество'] * 33
    pipe_parsed['T_23_2m100_50mm'] = pipe_parsed['mati_2m100_50mm * Количество'] * 36
    pipe_parsed['T_23_m100_70mm'] = pipe_parsed['mati_m100_70mm * Количество'] * 30
    pipe_parsed['T_23_m100_60mm'] = pipe_parsed['mati_m100_60mm * Количество'] * 33
    pipe_parsed['T_23_m100_50mm'] = pipe_parsed['mati_m100_50mm * Количество'] * 36

    ### Сопоставление толщины изоляции, диаметра трубы и типа (подающая/ надземная)
    ### Создаем новый датафрейм - объединенние парсера и крепежей для трубопроводов
    pipeline_with_insulation = pd.merge(pipe_parsed, pipeline_fasteners, on = ['Диаметр', 'δк'])
    pipeline_with_insulation['Лента 0,7х20 * Количество'] = pipeline_with_insulation['Количество'] * pipeline_with_insulation['Лента 0,7х20']
    pipeline_with_insulation['Лента 0,7х20 * Количество'] = pipeline_with_insulation['Лента 0,7х20 * Количество'].round(3)
    pipeline_with_insulation['Винт самонарезной * Количество'] = pipeline_with_insulation['Количество'] * pipeline_with_insulation['Винт самонарезной']
    pipeline_with_insulation['Винт самонарезной * Количество'] = pipeline_with_insulation['Винт самонарезной * Количество'].round(0)
    pipeline_with_insulation['Проволока 2-0-Ч * Количество'] = pipeline_with_insulation['Количество'] * pipeline_with_insulation['Проволока 2-0-Ч']
    pipeline_with_insulation['Проволока 2-0-Ч * Количество'] = pipeline_with_insulation['Проволока 2-0-Ч * Количество'].round(3)
    pipeline_with_insulation['Проволока 0,8-0-Ч * Количество'] = pipeline_with_insulation['Количество'] * pipeline_with_insulation['Проволока 0,8-0-Ч']
    pipeline_with_insulation['Проволока 0,8-0-Ч * Количество'] = pipeline_with_insulation['Проволока 0,8-0-Ч * Количество'].round(3)
    pipeline_with_insulation['Скоба опорная * Количество'] = pipeline_with_insulation['Количество'] * pipeline_with_insulation['Скоба опорная']
    pipeline_with_insulation['Скоба опорная * Количество'] = pipeline_with_insulation['Скоба опорная * Количество'].round(0)
    pipeline_with_insulation['Пряжка тип I-О * Количество'] = pipeline_with_insulation['Количество'] * pipeline_with_insulation['Пряжка тип I-О']
    pipeline_with_insulation['Пряжка тип I-О * Количество'] = pipeline_with_insulation['Пряжка тип I-О * Количество'].round(0)
    pipeline_with_insulation['Лента 2х30 * Длина вертикальных участков'] = pipeline_with_insulation['Лента 2х30'] * \
                                                                           pipeline_with_insulation['Длина вертикальных участков'] / 10
    pipeline_with_insulation['Лента 2х30 * Длина вертикальных участков'] = pipeline_with_insulation['Лента 2х30 * Длина вертикальных участков'].round(3)
    pipeline_with_insulation['Лента 3х30 * Длина вертикальных участков'] = pipeline_with_insulation['Лента 3х30'] * \
                                                                           pipeline_with_insulation['Длина вертикальных участков'] / 10
    pipeline_with_insulation['Лента 3х30 * Длина вертикальных участков'] = pipeline_with_insulation['Лента 3х30 * Длина вертикальных участков'].round(3)
    pipeline_with_insulation['Уголок 30х30x3 * Длина вертикальных участков'] = pipeline_with_insulation['Уголок 30х30x3'] * \
                                                                               pipeline_with_insulation['Длина вертикальных участков'] / 10
    pipeline_with_insulation['Уголок 30х30x3 * Длина вертикальных участков'] = pipeline_with_insulation['Уголок 30х30x3 * Длина вертикальных участков'].round(3)
    pipeline_with_insulation['Скоба навесная * Длина вертикальных участков'] = pipeline_with_insulation['Скоба навесная'] * \
                                                                               pipeline_with_insulation['Длина вертикальных участков'] / 10
    pipeline_with_insulation['Скоба навесная * Длина вертикальных участков'] = pipeline_with_insulation['Скоба навесная * Длина вертикальных участков'].round(3)
    pipeline_with_insulation['Лист АД1.Н-0.8 * Длина вертикальных участков'] = pipeline_with_insulation['Лист АД1.Н-0.8'] * \
                                                                               pipeline_with_insulation['Длина вертикальных участков'] / 10
    pipeline_with_insulation['Лист АД1.Н-0.8 * Длина вертикальных участков'] = pipeline_with_insulation['Лист АД1.Н-0.8 * Длина вертикальных участков'].round(3)
    pipeline_with_insulation['Болт М8 * Длина вертикальных участков'] = pipeline_with_insulation['Болт М8'] * \
                                                                        pipeline_with_insulation['Длина вертикальных участков'] / 10
    pipeline_with_insulation['Болт М8 * Длина вертикальных участков'] = pipeline_with_insulation['Болт М8 * Длина вертикальных участков'].round(3)
    pipeline_with_insulation['Болт М12 * Длина вертикальных участков'] = pipeline_with_insulation['Болт М12'] * \
                                                                         pipeline_with_insulation['Длина вертикальных участков'] / 10
    pipeline_with_insulation['Болт М12 * Длина вертикальных участков'] = pipeline_with_insulation['Болт М12 * Длина вертикальных участков'].round(3)
    pipeline_with_insulation['Гайка М8 * Длина вертикальных участков'] = pipeline_with_insulation['Гайка М8'] * \
                                                                         pipeline_with_insulation['Длина вертикальных участков'] / 10
    pipeline_with_insulation['Гайка М8 * Длина вертикальных участков'] = pipeline_with_insulation['Гайка М8 * Длина вертикальных участков'].round(3)
    pipeline_with_insulation['Гайка М12 * Длина вертикальных участков'] = pipeline_with_insulation['Гайка М12'] * \
                                                                          pipeline_with_insulation['Длина вертикальных участков'] / 10
    pipeline_with_insulation['Гайка М12 * Длина вертикальных участков'] = pipeline_with_insulation['Гайка М12 * Длина вертикальных участков'].round(3)
    ### Создаем колонки со значениями объемов для доп.крепежа вертикальных участков для мерзких сметчиков
    pipeline_with_insulation['mati_2m100_80mm * Длина вертикальных участков'] = pipeline_with_insulation[
                                                                                    'Длина вертикальных участков'] * \
                                                                                pipeline_with_insulation[
                                                                                    'mati_2m100_80mm']
    pipeline_with_insulation['mati_2m100_80mm * Длина вертикальных участков'] = pipeline_with_insulation[
        'mati_2m100_80mm * Длина вертикальных участков'].round(3)
    pipeline_with_insulation['mati_2m100_70mm * Длина вертикальных участков'] = pipeline_with_insulation[
                                                                                    'Длина вертикальных участков'] * \
                                                                                pipeline_with_insulation[
                                                                                    'mati_2m100_70mm']
    pipeline_with_insulation['mati_2m100_70mm * Длина вертикальных участков'] = pipeline_with_insulation[
        'mati_2m100_70mm * Длина вертикальных участков'].round(3)
    pipeline_with_insulation['mati_2m100_60mm * Длина вертикальных участков'] = pipeline_with_insulation[
                                                                                    'Длина вертикальных участков'] * \
                                                                                pipeline_with_insulation[
                                                                                    'mati_2m100_60mm']
    pipeline_with_insulation['mati_2m100_60mm * Длина вертикальных участков'] = pipeline_with_insulation[
        'mati_2m100_60mm * Длина вертикальных участков'].round(3)
    pipeline_with_insulation['mati_2m100_50mm * Длина вертикальных участков'] = pipeline_with_insulation[
                                                                                    'Длина вертикальных участков'] * \
                                                                                pipeline_with_insulation[
                                                                                    'mati_2m100_50mm']
    pipeline_with_insulation['mati_2m100_50mm * Длина вертикальных участков'] = pipeline_with_insulation[
        'mati_2m100_50mm * Длина вертикальных участков'].round(3)
    pipeline_with_insulation['mati_m100_70mm * Длина вертикальных участков'] = pipeline_with_insulation[
                                                                                   'Длина вертикальных участков'] * \
                                                                               pipeline_with_insulation[
                                                                                   'mati_m100_70mm']
    pipeline_with_insulation['mati_m100_70mm * Длина вертикальных участков'] = pipeline_with_insulation[
        'mati_m100_70mm * Длина вертикальных участков'].round(3)
    pipeline_with_insulation['mati_m100_60mm * Длина вертикальных участков'] = pipeline_with_insulation[
                                                                                   'Длина вертикальных участков'] * \
                                                                               pipeline_with_insulation[
                                                                                   'mati_m100_60mm']
    pipeline_with_insulation['mati_m100_60mm * Длина вертикальных участков'] = pipeline_with_insulation[
        'mati_m100_60mm * Длина вертикальных участков'].round(3)
    pipeline_with_insulation['mati_m100_50mm * Длина вертикальных участков'] = pipeline_with_insulation[
                                                                                   'Длина вертикальных участков'] * \
                                                                               pipeline_with_insulation[
                                                                                   'mati_m100_50mm']
    pipeline_with_insulation['mati_m100_50mm * Длина вертикальных участков'] = pipeline_with_insulation[
        'mati_m100_50mm * Длина вертикальных участков'].round(3)

    pipeline_with_insulation['Лак БТ-7,кг * Количество'] = np.nan
    pipeline_with_insulation['Лак БТ-7,м2 * Количество'] = np.nan
    for i in range(len(pipeline_with_insulation)):
        if pipeline_with_insulation['Тип прокладки'][i] == 'надземная':
            pipeline_with_insulation['Лак БТ-7,кг * Количество'][i] = 0
            pipeline_with_insulation['Лак БТ-7,м2 * Количество'][i] = 0
        else:
            pipeline_with_insulation['Лак БТ-7,кг * Количество'][i] = pipeline_with_insulation['Количество'][i] * pipeline_with_insulation['Лак БТ-7,кг'][i]
            pipeline_with_insulation['Лак БТ-7,кг * Количество'][i] = pipeline_with_insulation['Лак БТ-7,кг * Количество'][i].round(3)
            pipeline_with_insulation['Лак БТ-7,м2 * Количество'][i] = pipeline_with_insulation['Лента 0,7х20 * Количество'][i] / 5.495
            pipeline_with_insulation['Лак БТ-7,м2 * Количество'][i] = pipeline_with_insulation['Лак БТ-7,м2 * Количество'][i].round(3)
            pipeline_with_insulation['Винт самонарезной * Количество'][i] = 0


    pipeline_with_insulation = (pipeline_with_insulation.sort_values(by=['Код KKS', 'Тип прокладки', 'Диаметр'], ascending=[True, True, False])).reset_index()
    ### Сортировка по двум сразу, по возрастанию и убыванию
    del pipeline_with_insulation['index']
    #pipeline_with_insulation.to_excel("pipelines.xlsx", index=False)
    return pipeline_with_insulation

#pipeline_with_insulation = pipeline_insun_preporation()
### расчет арматуры
@snoop
def valves_insulation():
    if armatura_presence != "да":
        return None
    else:
        valves_fasteners = pd.merge(armatura_parsed, armatura_fasteners, on = ['Диаметр', 'Тип прокладки'])
        valves_fasteners['Лента АД1.08х20 * Количество'] = valves_fasteners['Лента АД1.08х20'] * valves_fasteners['Количество']
        valves_fasteners['Лента АД1.08х20 * Количество'] = valves_fasteners['Лента АД1.08х20 * Количество'].round(3)
        valves_fasteners['Проволока 0,8-0-Ч * Количество'] = valves_fasteners['Проволока 0,8-0-Ч'] * valves_fasteners['Количество']
        valves_fasteners['Проволока 0,8-0-Ч * Количество'] = valves_fasteners['Проволока 0,8-0-Ч * Количество'].round(3)
        valves_fasteners['Пряжка тип II-А * Количество'] = valves_fasteners['Пряжка тип II-А'] * valves_fasteners['Количество']
        valves_fasteners['Заклепка комбинированная ЗК-12-4,5 * Количество'] = valves_fasteners['Заклепка комбинированная ЗК-12-4,5'] * valves_fasteners['Количество']
        valves_fasteners['Крючок * Количество'] = valves_fasteners['Крючок'] * valves_fasteners['Количество']
        valves_fasteners['Крючок * Количество'] = valves_fasteners['Крючок * Количество'].round(3)
        valves_fasteners['Серьга * Количество'] = valves_fasteners['Серьга'] * valves_fasteners['Количество']
        valves_fasteners['Серьга * Количество'] = valves_fasteners['Серьга * Количество'].round(3)
        valves_fasteners['Рычаг * Количество'] = valves_fasteners['Рычаг'] * valves_fasteners['Количество']
        valves_fasteners['Рычаг * Количество'] = valves_fasteners['Рычаг * Количество'].round(3)
        valves_fasteners['Основание * Количество'] = valves_fasteners['Основание'] * valves_fasteners['Количество']
        valves_fasteners['Основание * Количество'] = valves_fasteners['Основание * Количество'].round(3)
        valves_fasteners['Заклепка 4x24.37 * Количество'] = valves_fasteners['Заклепка 4x24.37'] * valves_fasteners['Количество']
        valves_fasteners['Заклепка 4x24.37 * Количество'] = valves_fasteners['Заклепка 4x24.37 * Количество'].round(3)
        valves_insulation = pd.merge(valves_fasteners, armatura_pos1_5, on = ['Диаметр', 'Тип прокладки', 'δк'])
        valves_insulation["mati_2m100_60mm"] = np.nan
        valves_insulation["mati_2m100_40mm"] = np.nan
        valves_insulation["steel_0_8mm"] = np.nan
        valves_insulation['T_23_2m100_60mm'] = np.nan
        valves_insulation['T_23_2m100_40mm'] = np.nan
        valves_insulation['steel_0_8mm * Количество'] = np.nan
        for i in range(len(valves_insulation)):
            if valves_insulation['Тип прокладки'][i] == 'надземная':
                valves_insulation["mati_2m100_60mm"][i] = valves_insulation['V, м3'][i]
                valves_insulation['T_23_2m100_60mm'][i] = valves_insulation["mati_2m100_60mm"][i] * 33 * valves_insulation["Количество"][i]
                valves_insulation['T_23_2m100_60mm'][i] = valves_insulation['T_23_2m100_60mm'][i].round(3)
                valves_insulation["steel_0_8mm"][i] = valves_insulation['S, м2'][i]

            else:
                valves_insulation["mati_2m100_40mm"][i] = valves_insulation['V, м3'][i]
                valves_insulation['T_23_2m100_40mm'][i] = valves_insulation["mati_2m100_40mm"][i] * 43 * valves_insulation["Количество"][i]
                valves_insulation['T_23_2m100_40mm'][i] = valves_insulation['T_23_2m100_40mm'][i].round(3)
                valves_insulation["steel_0_8mm"][i] = valves_insulation['S, м2'][i]

        valves_insulation['mati_2m100_40mm * Количество'] = valves_insulation["mati_2m100_40mm"] * valves_insulation["Количество"]
        valves_insulation['mati_2m100_40mm * Количество'] = valves_insulation["mati_2m100_40mm * Количество"].round(3)
        valves_insulation['mati_2m100_60mm * Количество'] = valves_insulation["mati_2m100_60mm"] * valves_insulation["Количество"]
        valves_insulation['mati_2m100_60mm * Количество'] = valves_insulation['mati_2m100_60mm * Количество'].round(3)
        valves_insulation['steel_0_8mm * Количество'] = valves_insulation["steel_0_8mm"] * valves_insulation["Количество"]
        valves_insulation['steel_0_8mm * Количество'] = valves_insulation["steel_0_8mm * Количество"].round(3)
        valves_insulation = (valves_insulation.sort_values(by=['Код KKS', 'Тип прокладки', 'Диаметр' ], ascending=[True, True, False])).reset_index()
        #valves_insulation.to_excel("valves.xlsx", index=False)

        return valves_insulation

#valves_insulation = valves_insulation()
    ### Подготовка датафрейма для выводов
@snoop
def mpd_preporation():
    mpd_pipeline_with_insulation = pd.DataFrame(data = pipeline_with_insulation['Код KKS'])
    mpd_pipeline_with_insulation.insert(1,"Диаметр", pipeline_with_insulation['Диаметр'])
    mpd_pipeline_with_insulation.insert(2,"Тип прокладки", pipeline_with_insulation['Тип прокладки'])
    mpd_pipeline_with_insulation.insert(3,"Количество", pipeline_with_insulation['Количество'])
    mpd_pipeline_with_insulation.insert(4,"Наименование и техническая характеристика", pipeline_with_insulation['Наименование и техническая характеристика'])
    mpd_pipeline_with_insulation.insert(5,"Маты 2М-100 80мм", pipeline_with_insulation["mati_2m100_80mm"])
    mpd_pipeline_with_insulation.insert(6,"Маты 2М-100 70мм", pipeline_with_insulation["mati_2m100_70mm"])
    mpd_pipeline_with_insulation.insert(7,"Маты 2М-100 60мм", pipeline_with_insulation["mati_2m100_60mm"])
    mpd_pipeline_with_insulation.insert(8,"Маты 2М-100 50мм", pipeline_with_insulation["mati_2m100_50mm"])
    mpd_pipeline_with_insulation.insert(9,"Маты М-100 70мм", pipeline_with_insulation["mati_m100_70mm"])
    mpd_pipeline_with_insulation.insert(10, "Маты М-100 60мм", pipeline_with_insulation["mati_m100_60mm"])
    mpd_pipeline_with_insulation.insert(11, "Маты М-100 50мм", pipeline_with_insulation["mati_m100_50mm"])
    mpd_pipeline_with_insulation.insert(12, "Сталь 0.5мм", pipeline_with_insulation["steel_0_5mm"])
    mpd_pipeline_with_insulation.insert(13, "Сталь 0.8мм", pipeline_with_insulation["steel_0_8mm"])
    mpd_pipeline_with_insulation.insert(14, "Покрытие защ.Т-23", pipeline_with_insulation["protective_cover_T_23"])
    mpd_pipeline_with_insulation.insert(15, "δк", pipeline_with_insulation["δк"])
    mpd_pipeline_with_insulation.insert(16, "Температура", pipeline_with_insulation["Температура"])
    mpd_pipeline_with_insulation.insert(17, "Система", pipeline_with_insulation["Система"])
    mpd_pipeline_with_insulation = (mpd_pipeline_with_insulation.sort_values(by=['Код KKS', 'Тип прокладки', 'Диаметр'], ascending=[True, True, False])).reset_index()

    if armatura_presence == "да":
        mpd_valves_insulation = pd.DataFrame(data = valves_insulation['Код KKS'])
        mpd_valves_insulation.insert(1,"Диаметр", valves_insulation['Диаметр'])
        mpd_valves_insulation.insert(2,"Тип прокладки", valves_insulation['Тип прокладки'])
        mpd_valves_insulation.insert(3,"Количество", valves_insulation['Количество'])
        mpd_valves_insulation['Количество'] = mpd_valves_insulation['Количество'].round(0)
        mpd_valves_insulation.insert(4,"Наименование и техническая характеристика", valves_insulation['Наименование и техническая характеристика'])
        mpd_valves_insulation.insert(5,"Маты 2М-100 60мм", valves_insulation["mati_2m100_60mm"])
        mpd_valves_insulation.insert(6,"Маты 2М-100 40мм", valves_insulation["mati_2m100_40mm"])
        mpd_valves_insulation.insert(7,"Сталь 0.8мм", valves_insulation["steel_0_8mm"])
        mpd_valves_insulation.insert(8,"δк", valves_insulation["δк"])
        mpd_valves_insulation.insert(9,"Температура",  armatura_parsed["Температура"])
        mpd_valves_insulation.insert(10,"Система",  armatura_parsed["Система"])
        mpd = pd.concat([mpd_pipeline_with_insulation, mpd_valves_insulation], ignore_index=True)
    else:
        mpd = mpd_pipeline_with_insulation
    mpd = (mpd.sort_values(by=['Код KKS', 'Тип прокладки', 'Диаметр'], ascending=[True, True, False])).reset_index()
    #mpd = (mpd.sort_values(by=['Тип прокладки'], ascending=[True])).reset_index()
    # del mpd['index']
    #mpd.to_excel("mpd.xlsx", index=False)
    mpd_pipeline_with_insulation.to_excel("mpd_pipeline_with_insulation.xlsx", index=False)
    return mpd

#mpd = mpd_preporation()

### Подготовка списка с расходами ТИ на погонный метр труб, Колонки от 2М-100_80мм до М-100_50мм
@snoop
def insulation_list_prep():
    column_list = list(mpd.columns)
    slise_start = column_list.index('Наименование и техническая характеристика')
    slise_finish = column_list.index('Сталь 0.5мм')
    mpd_slise = pd.DataFrame(data = mpd.iloc[:,slise_start+1:slise_finish])
    mpd_slise = mpd_slise.fillna(0)
    #mpd_slise.to_excel("mpd_slise.xlsx", index=False)
    list_of_insul = []
    list_of_insulation_thickness = mpd_slise.values.tolist()
    for i in list_of_insulation_thickness:
        for j in i:
            if j != 0:
                list_of_insul.append(j)
    return list_of_insul

#list_of_insulation = insulation_list_prep()

### Делаем вывод
### Вывод в ведомость MPD0001
### Определяем диапазон для материала (маты). Выбираем откуда брать значения для записи объемов

#os.mkdir(КодKKS) - создаем директорию(папку) с названием = код KKS
@snoop

def output():

    def mpd_write():
        book = openpyxl.load_workbook("input/Templates/Template-MPD0001.xlsx")
        sheet = book['Sheet1']
        l = len(mpd) * 6
        i = 6
        i_df = 1
        j = 0
        n = 0 # для итерации по списку изоляции
        k = 5
        column_a = sheet['A']
        column_b = sheet['B']
        column_d = sheet['D']
        column_c = sheet['C']
        column_e = sheet['E']
        column_f = sheet['F']
        column_g = sheet['G']
        column_h = sheet['H']
        column_i = sheet['I']
        column_j = sheet['J']
        column_k = sheet['K']
        column_l = sheet['L']
        column_m = sheet['M']
        sheet.merge_cells(start_row=6, start_column=1, end_row=6, end_column=13)
        column_a[k].value = str(mpd['Код KKS'][0]) + ' ' + str(mpd['Тип прокладки'][0]) + ' прокладка''\n'+ str(mpd['Система'][0])
        column_a[k].alignment = Alignment(wrapText=True, horizontal="center")
        column_a[k].font = Font(name='Times New Roman', size=18, bold=True)
        sheet.row_dimensions[k+1].height = 50
        previous_KKS = mpd['Код KKS'][0]
        previous_ground_type = mpd['Тип прокладки'][0]
        while i <= l:
            if j > 0 and previous_KKS != mpd['Код KKS'][j]:
                previous_KKS = mpd['Код KKS'][j]
                column_a[k].value = mpd['Код KKS'][j] + ' ' + mpd['Тип прокладки'][j] + ' прокладка''\n' + mpd['Система'][j]
                column_a[k].alignment = Alignment(wrapText=True, horizontal="center")
                sheet.row_dimensions[k+1].height = 50
                column_a[k].font = Font(name='Times New Roman', size=18, bold=True)
                sheet.merge_cells(start_row=(k + 1), start_column=1, end_row=(k + 1), end_column=13)
            if j > 0 and previous_ground_type != mpd['Тип прокладки'][j]:
                previous_ground_type = mpd['Тип прокладки'][j]
                column_a[k].value = mpd['Код KKS'][j] + ' ' + mpd['Тип прокладки'][j] + ' прокладка'"\n" + mpd['Система'][j]
                column_a[k].alignment = Alignment(wrapText=True, horizontal="center")
                sheet.row_dimensions[k+1].height = 50
                column_a[k].font = Font(name='Times New Roman', size=18, bold=True)
                sheet.merge_cells(start_row=(k + 1), start_column=1, end_row=(k + 1), end_column=13)
            column_a[i].value = i_df
            column_a[i].number_format = "0"
            if 'Кран' in mpd['Наименование и техническая характеристика'][j]:
                column_b[i].value = 'Кран шаровой' + ' DN' + str(mpd['Диаметр'][j])
                column_b[i].alignment = Alignment(horizontal="left")
                if mpd['Диаметр'][j] < 250:
                    column_m[i].value = "7.903.9-8.15.3-07"
                    column_m[i].alignment = Alignment(horizontal="center")
                    column_m[i + 1].value = "(применительно)"
                    column_m[i + 1].alignment = Alignment(horizontal="center")
                elif 250 <= mpd['Диаметр'][j] <= 600:
                    column_m[i].value = "7.903.9-8.15.3-09"
                    column_m[i].alignment = Alignment(horizontal="center")
                    column_m[i + 1].value = "(применительно)"
                    column_m[i + 1].alignment = Alignment(horizontal="center")
                else:
                    column_m[i].value = "7.903.9-8.15.3-11"
                    column_m[i].alignment = Alignment(horizontal="center")
                    column_m[i + 1].value = "(применительно)"
                    column_m[i + 1].alignment = Alignment(horizontal="center")
            elif 'адвижка' in mpd['Наименование и техническая характеристика'][j]:
                column_b[i].value = 'Задвижка клиновая' + ' DN' + str(mpd['Диаметр'][j])
                column_b[i].alignment = Alignment(horizontal="left")
                if mpd['Диаметр'][j] < 250:
                    column_m[i].value = "7.903.9-8.15.3-07"
                    column_m[i].alignment = Alignment(horizontal="center")
                    column_m[i + 1].value = "(применительно)"
                    column_m[i + 1].alignment = Alignment(horizontal="center")
                elif 250 <= mpd['Диаметр'][j] <= 600:
                    column_m[i].value = "7.903.9-8.15.3-09"
                    column_m[i].alignment = Alignment(horizontal="center")
                    column_m[i + 1].value = "(применительно)"
                    column_m[i + 1].alignment = Alignment(horizontal="center")
                else:
                    column_m[i].value = "7.903.9-8.15.3-11"
                    column_m[i].alignment = Alignment(horizontal="center")
                    column_m[i + 1].value = "(применительно)"
                    column_m[i + 1].alignment = Alignment(horizontal="center")
            else:
                column_b[i].value = mpd['Наименование и техническая характеристика'][j]
                column_b[i].alignment = Alignment(horizontal="left")
                if mpd['Диаметр'][j] < 250:
                    column_m[i].value = "7.903.9-8.15.1-19"
                    column_m[i].alignment = Alignment(horizontal="center")
                    column_m[i+1].value = "(применительно)"
                    column_m[i + 1].alignment = Alignment(horizontal="center")
                elif 250 <= mpd['Диаметр'][j] <= 1420:
                    column_m[i].value = "7.903.9-8.15.1-14"
                    column_m[i].alignment = Alignment(horizontal="center")
                    column_m[i + 1].value = "(применительно)"
                    column_m[i + 1].alignment = Alignment(horizontal="center")

            if 600 <= mpd['Диаметр'][j] and mpd['Тип прокладки'][j] == 'надземная':
                column_f[i].value = 'Откр.'
                column_f[i + 1].value = 'возд.'
                column_e[i].value = mpd['Температура'][j]
                if 'Труба' in mpd['Наименование и техническая характеристика'][j]:
                    s = str(mpd['δк'][j] + 10)
                    column_c[i].value = 'м'
                    column_d[i].value = mpd['Количество'][j]
                    column_d[i].number_format = "0.00"
                    column_d[i].alignment = Alignment(horizontal="center")
                    column_g[i + 1].value = '2М-100 в обкладке из стеклоткани марки'
                    column_g[i + 1].alignment = Alignment(horizontal="left")
                    column_g[i + 2].value = 'марки Т-23 со всех сторон' + ' ' * 11 + 's = ' + s + ' мм'
                    column_g[i + 2].alignment = Alignment(horizontal="left")
                    column_g[i + 3].value = 'Сталь тонколистовая оцинкованная'
                    column_g[i + 3].alignment = Alignment(horizontal="left")
                    column_g[i + 4].value = 's = 0.8 мм'
                    column_g[i + 4].alignment = Alignment(horizontal="right")
                    column_h[i + 2].value = mpd['δк'][j]
                    column_i[i + 4].value = mpd['Сталь 0.8мм'][j]
                    column_i[i + 4].number_format = "0.000"
                    column_j[i + 4].value = (mpd['Сталь 0.8мм'][j] * mpd['Количество'][j]).round(3)
                    column_k[i + 2].value = list_of_insulation[n]
                    column_k[i + 2].number_format = "0.000"
                    column_l[i + 2].value = (column_k[i + 2].value * column_d[i].value).round(3)
                    column_l[i + 2].number_format = "0.000"
                    n += 1
                else:
                    s = str(mpd['δк'][j] + 10)
                    column_c[i].value = 'шт.'
                    column_d[i].value = mpd['Количество'][j]
                    column_d[i].number_format = "0"
                    column_d[i].alignment = Alignment(horizontal="center")
                    column_g[i + 1].value = '2М-100 в обкладке из стеклоткани марки'
                    column_g[i + 1].alignment = Alignment(horizontal="left")
                    column_g[i + 2].value = 'марки Т-23 со всех сторон' + ' ' * 11 + 's = ' + s + ' мм'
                    column_g[i + 2].alignment = Alignment(horizontal="left")
                    column_g[i + 3].value = 'Сталь тонколистовая оцинкованная'
                    column_g[i + 3].alignment = Alignment(horizontal="left")
                    column_g[i + 4].value = 's = 0.8 мм'
                    column_g[i + 4].alignment = Alignment(horizontal="right")
                    column_h[i + 2].value = mpd['δк'][j]
                    column_i[i + 4].value = mpd['Сталь 0.8мм'][j]
                    column_i[i + 4].number_format = "0.000"
                    column_j[i + 4].value = mpd['Сталь 0.8мм'][j] * mpd['Количество'][j]
                    column_j[i + 4].number_format = "0.000"
                    column_k[i + 2].value = mpd['Маты 2М-100 60мм'][j]
                    column_k[i + 2].number_format = "0.000"
                    column_l[i + 2].value = mpd['Маты 2М-100 60мм'][j] * mpd['Количество'][j]
                    column_l[i + 2].number_format = "0.000"
                    n += 1

            elif 250 <= mpd['Диаметр'][j] < 600 and mpd['Тип прокладки'][j] == 'надземная':
                column_f[i].value = 'Откр.'
                column_f[i + 1].value = 'возд.'
                column_e[i].value = mpd['Температура'][j]
                if 'Труба' in mpd['Наименование и техническая характеристика'][j]:
                    s = str(mpd['δк'][j] + 10)
                    column_c[i].value = 'м'
                    column_d[i].value = mpd['Количество'][j]
                    column_d[i].number_format = "0.00"
                    column_d[i].alignment = Alignment(horizontal="center")
                    column_g[i + 1].value = '2М-100 в обкладке из стеклоткани марки'
                    column_g[i + 1].alignment = Alignment(horizontal="left")
                    column_g[i + 2].value = 'марки Т-23 со всех сторон' + ' ' * 11 + 's = ' + s + ' мм'
                    column_g[i + 2].alignment = Alignment(horizontal="left")
                    column_g[i + 3].value = 'Сталь тонколистовая оцинкованная'
                    column_g[i + 3].alignment = Alignment(horizontal="left")
                    column_g[i + 4].value = 's = 0.5 мм'
                    column_g[i + 4].alignment = Alignment(horizontal="right")
                    column_h[i + 2].value = mpd['δк'][j]
                    column_i[i + 4].value = mpd['Сталь 0.5мм'][j]
                    column_i[i + 4].number_format = "0.000"
                    column_j[i + 4].value = (mpd['Сталь 0.5мм'][j] * mpd['Количество'][j]).round(3)
                    column_k[i + 2].value = list_of_insulation[n]
                    column_k[i + 2].number_format = "0.000"
                    column_l[i + 2].value = (column_k[i + 2].value * column_d[i].value).round(3)
                    column_l[i + 2].number_format = "0.000"
                    n += 1
                else:
                    s = str(mpd['δк'][j] + 10)
                    column_c[i].value = 'шт.'
                    column_d[i].value = mpd['Количество'][j]
                    column_d[i].number_format = "0"
                    column_d[i].alignment = Alignment(horizontal="center")
                    column_g[i + 1].value = '2М-100 в обкладке из стеклоткани марки'
                    column_g[i + 1].alignment = Alignment(horizontal="left")
                    column_g[i + 2].value = 'марки Т-23 со всех сторон' + ' ' * 11 + 's = ' + s + ' мм'
                    column_g[i + 2].alignment = Alignment(horizontal="left")
                    column_g[i + 3].value = 'Сталь тонколистовая оцинкованная'
                    column_g[i + 3].alignment = Alignment(horizontal="left")
                    column_g[i + 4].value = 's = 0.8 мм'
                    column_g[i + 4].alignment = Alignment(horizontal="right")
                    column_h[i + 2].value = mpd['δк'][j]
                    column_i[i + 4].value = mpd['Сталь 0.8мм'][j]
                    column_i[i + 4].number_format = "0.000"
                    column_j[i + 4].value = mpd['Сталь 0.8мм'][j] * mpd['Количество'][j]
                    column_j[i + 4].number_format = "0.000"
                    column_k[i + 2].value = mpd['Маты 2М-100 60мм'][j]
                    column_k[i + 2].number_format = "0.000"
                    column_l[i + 2].value = mpd['Маты 2М-100 60мм'][j] * mpd['Количество'][j]
                    column_l[i + 2].number_format = "0.000"
                    n += 1

            elif mpd['Диаметр'][j] < 250 and mpd['Тип прокладки'][j] == 'надземная':
                column_f[i].value = 'Откр.'
                column_f[i + 1].value = 'возд.'
                column_e[i].value = mpd['Температура'][j]
                if 'Труба' in mpd['Наименование и техническая характеристика'][j]:
                    s = str(mpd['δк'][j] + 10)
                    column_c[i].value = 'м'
                    column_d[i].value = mpd['Количество'][j]
                    column_d[i].number_format = "0.00"
                    column_d[i].alignment = Alignment(horizontal="center")
                    column_g[i + 1].value = 'М-100 без обкладок'
                    column_g[i + 1].alignment = Alignment(horizontal="left")
                    column_g[i + 2].value = 's = ' + s + ' мм'
                    column_g[i + 2].alignment = Alignment(horizontal="right")
                    column_g[i + 3].value = 'Сталь тонколистовая оцинкованная'
                    column_g[i + 3].alignment = Alignment(horizontal="left")
                    column_g[i + 4].value = 's = 0.5 мм'
                    column_g[i + 4].alignment = Alignment(horizontal="right")
                    column_h[i + 2].value = mpd['δк'][j]
                    column_h[i + 2].number_format = "0"
                    column_i[i + 4].value = mpd['Сталь 0.5мм'][j]
                    column_i[i + 4].alignment = Alignment(horizontal="center")
                    column_i[i + 4].number_format = "0.000"
                    column_j[i + 4].value = (mpd['Сталь 0.5мм'][j] * mpd['Количество'][j]).round(3)
                    column_j[i + 4].number_format = "0.000"
                    column_j[i + 4].alignment = Alignment(horizontal="center")
                    column_k[i + 2].value = list_of_insulation[n]
                    column_k[i + 2].number_format = "0.000"
                    column_k[i + 2].alignment = Alignment(horizontal="center")
                    column_l[i + 2].value = (column_k[i + 2].value * column_d[i].value).round(3)
                    column_l[i + 2].number_format = "0.000"
                    column_l[i + 2].alignment = Alignment(horizontal="center")
                    n += 1
                else:
                    s = str(mpd['δк'][j] + 10)
                    column_c[i].value = 'шт.'
                    column_d[i].value = mpd['Количество'][j]
                    column_d[i].number_format = "0"
                    column_d[i].alignment = Alignment(horizontal="center")
                    column_g[i + 1].value = '2М-100 в обкладке из стеклоткани марки'
                    column_g[i + 1].alignment = Alignment(horizontal="left")
                    column_g[i + 2].value = 'марки Т-23 со всех сторон' + ' ' * 11 + 's = ' + s + ' мм'
                    column_g[i + 2].alignment = Alignment(horizontal="left")
                    column_g[i + 3].value = 'Сталь тонколистовая оцинкованная'
                    column_g[i + 3].alignment = Alignment(horizontal="left")
                    column_g[i + 4].value = 's = 0.8 мм'
                    column_g[i + 4].alignment = Alignment(horizontal="right")
                    column_h[i + 2].value = mpd['δк'][j]
                    column_i[i + 4].value = mpd['Сталь 0.8мм'][j]
                    column_i[i + 4].number_format = "0.000"
                    column_j[i + 4].value = mpd['Сталь 0.8мм'][j] * mpd['Количество'][j]
                    column_j[i + 4].number_format = "0.000"
                    column_k[i + 2].value = mpd['Маты 2М-100 60мм'][j]
                    column_k[i + 2].number_format = "0.000"
                    column_l[i + 2].value = mpd['Маты 2М-100 60мм'][j] * mpd['Количество'][j]
                    column_l[i + 2].number_format = "0.000"
                    n += 1

            if mpd['Диаметр'][j] > 400 and mpd['Тип прокладки'][j] == 'подземная':
                column_f[i].value = 'Непр.'
                column_f[i + 1].value = 'канал'
                column_e[i].value = mpd['Температура'][j]
                if 'Труба' in mpd['Наименование и техническая характеристика'][j]:
                    s = str(mpd['δк'][j] + 10)
                    column_c[i].value = 'м'
                    column_d[i].value = mpd['Количество'][j]
                    column_d[i].number_format = "0.000"
                    column_d[i].alignment = Alignment(horizontal="center")
                    column_g[i + 1].value = '2М-100 в обкладке из стеклоткани марки'
                    column_g[i + 1].alignment = Alignment(horizontal="left")
                    column_g[i + 2].value = 'марки Т-23 со всех сторон' + ' ' * 11 + 's = ' + s + ' мм'
                    column_g[i + 2].alignment = Alignment(horizontal="left")
                    column_g[i + 3].value = 'Покрытие защитное из стеклоткани'
                    column_g[i + 3].alignment = Alignment(horizontal="left")
                    column_g[i + 4].value = 'марки T-23'
                    column_g[i + 4].alignment = Alignment(horizontal="left")
                    column_h[i + 2].value = mpd['δк'][j]
                    column_i[i + 4].value = mpd['Покрытие защ.Т-23'][j]
                    column_i[i + 4].number_format = "0.000"
                    column_j[i + 4].value = (mpd['Покрытие защ.Т-23'][j] * mpd['Количество'][j]).round(3)
                    column_j[i + 4].number_format = "0.000"
                    column_k[i + 2].value = mpd['Маты 2М-100 70мм'][j]
                    column_k[i + 2].number_format = "0.000"
                    column_l[i + 2].value = (column_k[i + 2].value * column_d[i].value).round(3)
                    column_l[i + 2].number_format = "0.000"
                    n += 1
                else:
                    s = str(mpd['δк'][j] + 10)
                    column_c[i].value = 'шт.'
                    column_d[i].value = mpd['Количество'][j]
                    column_d[i].number_format = "0"
                    column_d[i].alignment = Alignment(horizontal="center")
                    column_g[i + 1].value = '2М-100 в обкладках из стеклоткани'
                    column_g[i + 1].alignment = Alignment(horizontal="left")
                    column_g[i + 2].value = 'марки Т-23 со всех сторон' + ' ' * 11 + 's = ' + s + ' мм'
                    column_g[i + 3].value = 'Сталь тонколистовая оцинкованная'
                    column_g[i + 3].alignment = Alignment(horizontal="left")
                    column_g[i + 4].value = 's = 0.8 мм'
                    column_g[i + 4].alignment = Alignment(horizontal="right")
                    column_h[i + 2].value = mpd['δк'][j]
                    column_i[i + 4].value = mpd['Сталь 0.8мм'][j]
                    column_i[i + 4].number_format = "0.000"
                    column_j[i + 4].value = mpd['Сталь 0.8мм'][j] * mpd['Количество'][j]
                    column_j[i + 4].number_format = "0.000"
                    column_k[i + 2].value = mpd['Маты 2М-100 40мм'][j]
                    column_k[i + 2].number_format = "0.000"
                    column_l[i + 2].value = mpd['Маты 2М-100 40мм'][j] * mpd['Количество'][j]
                    column_l[i + 2].number_format = "0.000"
                    n += 1

            if 250 <= mpd['Диаметр'][j] <= 400 and mpd['Тип прокладки'][j] == 'подземная':
                column_f[i].value = 'Непр.'
                column_f[i + 1].value = 'канал'
                column_e[i].value = mpd['Температура'][j]
                if 'Труба' in mpd['Наименование и техническая характеристика'][j]:
                    s = str(mpd['δк'][j] + 10)
                    column_c[i].value = 'м'
                    column_d[i].value = mpd['Количество'][j]
                    column_d[i].number_format = "0.000"
                    column_d[i].alignment = Alignment(horizontal="center")
                    column_g[i + 1].value = '2М-100 в обкладке из стеклоткани марки'
                    column_g[i + 1].alignment = Alignment(horizontal="left")
                    column_g[i + 2].value = 'марки Т-23 со всех сторон' + ' ' * 11 + 's = ' + s + ' мм'
                    column_g[i + 2].alignment = Alignment(horizontal="left")
                    column_g[i + 3].value = 'Покрытие защитное из стеклоткани'
                    column_g[i + 3].alignment = Alignment(horizontal="left")
                    column_g[i + 4].value = 'марки T-23'
                    column_g[i + 4].alignment = Alignment(horizontal="left")
                    column_h[i + 2].value = mpd['δк'][j]
                    column_i[i + 4].value = mpd['Покрытие защ.Т-23'][j]
                    column_i[i + 4].number_format = "0.000"
                    column_j[i + 4].value = (mpd['Покрытие защ.Т-23'][j] * mpd['Количество'][j]).round(3)
                    column_j[i + 4].number_format = "0.000"
                    column_k[i + 2].value = mpd['Маты 2М-100 60мм'][j]
                    column_k[i + 2].number_format = "0.000"
                    column_l[i + 2].value = (column_k[i + 2].value * column_d[i].value).round(3)
                    column_l[i + 2].number_format = "0.000"
                    n += 1
                else:
                    s = str(mpd['δк'][j] + 10)
                    column_c[i].value = 'шт.'
                    column_d[i].value = mpd['Количество'][j]
                    column_d[i].number_format = "0"
                    column_g[i + 1].value = '2М-100 в обкладках из стеклоткани'
                    column_g[i + 1].alignment = Alignment(horizontal="left")
                    column_g[i + 2].value = 'марки Т-23 со всех сторон' + ' ' * 11 + 's = ' + s + ' мм'
                    column_g[i + 3].value = 'Сталь тонколистовая оцинкованная'
                    column_g[i + 3].alignment = Alignment(horizontal="left")
                    column_g[i + 4].value = 's = 0.8 мм'
                    #column_g[i + 5].value = '' ### костыль какой-то, разобраться откуда второе s вылазит
                    column_g[i + 4].alignment = Alignment(horizontal="right")
                    column_h[i + 2].value = mpd['δк'][j]
                    column_i[i + 4].value = mpd['Сталь 0.8мм'][j]
                    column_i[i + 4].number_format = "0.000"
                    column_j[i + 4].value = mpd['Сталь 0.8мм'][j] * mpd['Количество'][j]
                    column_j[i + 4].number_format = "0.000"
                    column_k[i + 2].value = mpd['Маты 2М-100 40мм'][j]
                    column_k[i + 2].number_format = "0.000"
                    column_l[i + 2].value = mpd['Маты 2М-100 40мм'][j] * mpd['Количество'][j]
                    column_l[i + 2].number_format = "0.000"
                    n += 1

            if 133 < mpd['Диаметр'][j] < 250 and mpd['Тип прокладки'][j] == 'подземная':
                column_f[i].value = 'Непр.'
                column_f[i + 1].value = 'канал'
                column_e[i].value = mpd['Температура'][j]
                if 'Труба' in mpd['Наименование и техническая характеристика'][j]:
                    s = str(mpd['δк'][j] + 10)
                    column_c[i].value = 'м'
                    column_d[i].value = mpd['Количество'][j]
                    column_d[i].number_format = "0.000"
                    column_d[i].alignment = Alignment(horizontal="center")
                    column_g[i + 1].value = 'М-100 без обкладок' + ' ' * 25 + 's = ' + s + ' мм'
                    column_g[i + 2].alignment = Alignment(horizontal="left")
                    column_g[i + 3].value = 'Покрытие защитное из стеклоткани'
                    column_g[i + 3].alignment = Alignment(horizontal="left")
                    column_g[i + 4].value = 'марки T-23'
                    column_g[i + 4].alignment = Alignment(horizontal="left")
                    column_h[i + 2].value = mpd['δк'][j]
                    column_i[i + 4].value = mpd['Покрытие защ.Т-23'][j]
                    column_i[i + 4].number_format = "0.000"
                    column_j[i + 4].value = (mpd['Покрытие защ.Т-23'][j] * mpd['Количество'][j]).round(3)
                    column_j[i + 4].number_format = "0.000"
                    column_k[i + 2].value = mpd['Маты М-100 60мм'][j]
                    column_k[i + 2].number_format = "0.000"
                    column_l[i + 2].value = (column_k[i + 2].value * column_d[i].value).round(3)
                    column_l[i + 2].number_format = "0.000"
                    n += 1
                else:
                    s = str(mpd['δк'][j] + 10)
                    column_c[i].value = 'шт.'
                    column_d[i].value = mpd['Количество'][j]
                    column_d[i].number_format = "0"
                    column_g[i + 1].value = '2М-100 в обкладках из стеклоткани'
                    column_g[i + 1].alignment = Alignment(horizontal="left")
                    column_g[i + 2].value = 'марки Т-23 со всех сторон' + ' ' * 11 + 's = ' + s + ' мм'
                    column_g[i + 2].alignment = Alignment(horizontal="left")
                    column_g[i + 3].value = 'Сталь тонколистовая оцинкованная'
                    column_g[i + 3].alignment = Alignment(horizontal="left")
                    column_g[i + 4].value = 's = 0.8 мм'
                    column_g[i + 4].alignment = Alignment(horizontal="right")
                    column_h[i + 2].value = mpd['δк'][j]
                    column_i[i + 4].value = mpd['Сталь 0.8мм'][j]
                    column_i[i + 4].number_format = "0.000"
                    column_j[i + 4].value = mpd['Сталь 0.8мм'][j] * mpd['Количество'][j]
                    column_j[i + 4].number_format = "0.000"
                    column_k[i + 2].value = mpd['Маты 2М-100 40мм'][j]
                    column_k[i + 2].number_format = "0.000"
                    column_l[i + 2].value = mpd['Маты 2М-100 40мм'][j] * mpd['Количество'][j]
                    column_l[i + 2].number_format = "0.000"
                    n += 1

            if mpd['Диаметр'][j] <= 133 and mpd['Тип прокладки'][j] == 'подземная':
                column_f[i].value = 'Непр.'
                column_f[i + 1].value = 'канал'
                column_e[i].value = mpd['Температура'][j]
                if 'Труба' in mpd['Наименование и техническая характеристика'][j]:
                    s = str(mpd['δк'][j] + 10)
                    column_c[i].value = 'м'
                    column_d[i].value = mpd['Количество'][j]
                    column_d[i].number_format = "0.000"
                    column_d[i].alignment = Alignment(horizontal="center")
                    column_g[i + 1].value = 'М-100 без обкладок' + ' ' * 25 + 's = ' + s + ' мм'
                    column_g[i + 1].alignment = Alignment(horizontal="left")
                    column_g[i + 2].value = 'Покрытие защитное из стеклоткани'
                    column_g[i + 2].alignment = Alignment(horizontal="left")
                    column_g[i + 3].value = 'марки T-23'
                    column_g[i + 3].alignment = Alignment(horizontal="left")
                    column_h[i + 2].value = mpd['δк'][j]
                    column_i[i + 4].value = mpd['Покрытие защ.Т-23'][j]
                    column_i[i + 4].number_format = "0.000"
                    column_j[i + 4].value = (mpd['Покрытие защ.Т-23'][j] * mpd['Количество'][j]).round(3)
                    column_j[i + 4].number_format = "0.000"
                    column_k[i + 2].value = mpd['Маты М-100 50мм'][j]
                    column_k[i + 2].number_format = "0.000"
                    column_k[i + 2].alignment = Alignment(horizontal="center")
                    column_l[i + 2].value = (column_k[i + 2].value * column_d[i].value).round(3)
                    column_l[i + 2].number_format = "0.000"
                    column_l[i + 2].alignment = Alignment(horizontal="center")
                    n += 1
                else:
                    s = str(mpd['δк'][j] + 10)
                    column_c[i].value = 'шт.'
                    column_d[i].value = mpd['Количество'][j]
                    column_d[i].number_format = "0"
                    column_g[i + 1].value = '2М-100 в обкладках из стеклоткани'
                    column_g[i + 1].alignment = Alignment(horizontal="left")
                    column_g[i + 2].value = 'марки Т-23 со всех сторон' + ' ' * 11 + 's = ' + s + ' мм'
                    column_g[i + 3].value = 'Сталь тонколистовая оцинкованная'
                    column_g[i + 3].alignment = Alignment(horizontal="left")
                    column_g[i + 4].value = 's = 0.8 мм'
                    column_g[i + 4].alignment = Alignment(horizontal="right")
                    column_h[i + 2].value = mpd['δк'][j]
                    column_i[i + 4].value = mpd['Сталь 0.8мм'][j]
                    column_i[i + 4].number_format = "0.000"
                    column_j[i + 4].value = mpd['Сталь 0.8мм'][j] * mpd['Количество'][j]
                    column_j[i + 4].number_format = "0.000"
                    column_k[i + 2].value = mpd['Маты 2М-100 40мм'][j]
                    column_k[i + 2].number_format = "0.000"
                    column_l[i + 2].value = mpd['Маты 2М-100 40мм'][j] * mpd['Количество'][j]
                    column_l[i + 2].number_format = "0.000"
                    n += 1

            column_g[i].value = 'Маты минераловатные прошивные марки'
            column_g[i].alignment = Alignment(horizontal="left")
            column_i[i + 2].value = '-'
            column_i[i + 2].alignment = Alignment(horizontal="center")
            column_j[i + 2].value = '-'
            column_j[i + 2].alignment = Alignment(horizontal="center")
            column_k[i + 4].value = '-'
            column_k[i + 4].alignment = Alignment(horizontal="center")
            column_l[i + 4].value = '-'
            column_l[i + 4].alignment = Alignment(horizontal="center")

            i += 6
            i_df += 1
            j += 1
            k += 6
        # sheet.delete_rows(6)
        ### Записываем колонтитул
        sheet.oddFooter.left.text = КодKKS + '-MPD0001' + '\n' + Архивный
        return book
    book = mpd_write()
    book.save(f"output/{КодKKS}-MPD0001.xlsx")

    ### Делаем вывод
    ### Вывод в ведомость MPA0001

    def mpa_write():
        # book = openpyxl.load_workbook("Template-MPA0001.xlsx")
        # sheet = book['Sheet1']
        mpa = pd.concat([pipeline_with_insulation, valves_insulation], ignore_index=True)
        mpa = (mpa.sort_values(by=['Код KKS', 'Тип прокладки', 'Диаметр'], ascending=[True, True, False])).reset_index()
        # mpa.to_excel("mpa.xlsx", index=False)
        if armatura_presence == "да":
            del mpa['steel_0_5mm']
            del mpa['Лента 0,7х20']
            del mpa['Винт самонарезной']
            del mpa['Проволока 2-0-Ч']
            del mpa['Проволока 0,8-0-Ч']
            del mpa['Скоба опорная']
            del mpa['Пряжка тип I-О']
            del mpa['Лак БТ-7,кг']
            del mpa['V, м3']
            del mpa['S, м2']
            del mpa['steel_0_8mm']
            del mpa['Лента АД1.08х20']
            del mpa['Пряжка тип II-А']
            del mpa['Заклепка комбинированная ЗК-12-4,5']
            del mpa['Наименование и техническая характеристика']
            del mpa['Примечание']
            del mpa['Крючок']
            del mpa['Серьга']
            del mpa['Рычаг']
            del mpa['Основание']
            del mpa['Заклепка 4x24.37']
            mpa_slise = mpa.groupby(['Код KKS', 'Тип прокладки'], as_index=False)['mati_2m100_80mm * Количество',
                                                                                  'mati_2m100_70mm * Количество', 'mati_2m100_60mm * Количество', 'mati_2m100_50mm * Количество',
                                                                                  'mati_2m100_40mm * Количество', 'mati_m100_70mm * Количество', 'mati_m100_60mm * Количество',
                                                                                  'mati_m100_50mm * Количество', 'protective_cover_T_23', 'steel_0_5mm * Количество', 'steel_0_8mm * Количество',
                                                                                  'T_23_2m100_80mm', 'T_23_2m100_70mm', 'T_23_2m100_60mm', 'T_23_2m100_50mm', 'T_23_m100_70mm', 'T_23_m100_60mm',
                                                                                  'T_23_m100_50mm', 'T_23_2m100_40mm', 'Лента 0,7х20 * Количество', 'Винт самонарезной * Количество',
                                                                                  'Проволока 2-0-Ч * Количество', 'Проволока 0,8-0-Ч * Количество', 'Скоба опорная * Количество',
                                                                                  'Пряжка тип I-О * Количество', 'Лак БТ-7,кг * Количество', 'Лак БТ-7,м2 * Количество', 'Лента АД1.08х20 * Количество',
                                                                                  'Пряжка тип II-А * Количество', 'Заклепка комбинированная ЗК-12-4,5 * Количество', 'Крючок * Количество',
                                                                                  'Серьга * Количество', 'Рычаг * Количество', 'Основание * Количество', 'Заклепка 4x24.37 * Количество',
                                                                                  'Лента 2х30 * Длина вертикальных участков', 'Лента 3х30 * Длина вертикальных участков', 'Уголок 30х30x3 * Длина вертикальных участков',
                                                                                  'Скоба навесная * Длина вертикальных участков', 'Лист АД1.Н-0.8 * Длина вертикальных участков', 'Болт М8 * Длина вертикальных участков',
                                                                                  'Болт М12 * Длина вертикальных участков', 'Гайка М8 * Длина вертикальных участков', 'Гайка М12 * Длина вертикальных участков',
                                                                                  'mati_2m100_80mm * Длина вертикальных участков', 'mati_2m100_70mm * Длина вертикальных участков',
                                                                                  'mati_2m100_60mm * Длина вертикальных участков', 'mati_2m100_50mm * Длина вертикальных участков',
                                                                                  'mati_m100_70mm * Длина вертикальных участков', 'mati_m100_60mm * Длина вертикальных участков',
                                                                                  'mati_m100_50mm * Длина вертикальных участков'].sum()
        else:
            del mpa['steel_0_5mm']
            del mpa['Лента 0,7х20']
            del mpa['Винт самонарезной']
            del mpa['Проволока 2-0-Ч']
            del mpa['Проволока 0,8-0-Ч']
            del mpa['Скоба опорная']
            del mpa['Пряжка тип I-О']
            del mpa['Лак БТ-7,кг']
            del mpa['steel_0_8mm']
            del mpa['Наименование и техническая характеристика']
            del mpa['Примечание']
            mpa_slise = mpa.groupby(['Код KKS', 'Тип прокладки'], as_index=False)['mati_2m100_80mm * Количество',
                                                                                  'mati_2m100_70mm * Количество', 'mati_2m100_60mm * Количество', 'mati_2m100_50mm * Количество',
                                                                                  'mati_m100_70mm * Количество', 'mati_m100_60mm * Количество',
                                                                                  'mati_m100_50mm * Количество', 'protective_cover_T_23', 'steel_0_5mm * Количество', 'steel_0_8mm * Количество',
                                                                                  'T_23_2m100_80mm', 'T_23_2m100_70mm', 'T_23_2m100_60mm', 'T_23_2m100_50mm', 'T_23_m100_70mm', 'T_23_m100_60mm',
                                                                                  'T_23_m100_50mm', 'Лента 0,7х20 * Количество', 'Винт самонарезной * Количество',
                                                                                  'Проволока 2-0-Ч * Количество', 'Проволока 0,8-0-Ч * Количество', 'Скоба опорная * Количество',
                                                                                  'Пряжка тип I-О * Количество', 'Лак БТ-7,кг * Количество', 'Лак БТ-7,м2 * Количество',
                                                                                  'Лента 2х30 * Длина вертикальных участков', 'Лента 3х30 * Длина вертикальных участков', 'Уголок 30х30x3 * Длина вертикальных участков',
                                                                                  'Скоба навесная * Длина вертикальных участков', 'Лист АД1.Н-0.8 * Длина вертикальных участков', 'Болт М8 * Длина вертикальных участков',
                                                                                  'Болт М12 * Длина вертикальных участков', 'Гайка М8 * Длина вертикальных участков', 'Гайка М12 * Длина вертикальных участков',
                                                                                  'mati_2m100_80mm * Длина вертикальных участков', 'mati_2m100_70mm * Длина вертикальных участков',
                                                                                  'mati_2m100_60mm * Длина вертикальных участков', 'mati_2m100_50mm * Длина вертикальных участков',
                                                                                  'mati_m100_70mm * Длина вертикальных участков', 'mati_m100_60mm * Длина вертикальных участков',
                                                                                  'mati_m100_50mm * Длина вертикальных участков'].sum()
        #mpa.to_excel("mpa.xlsx", index=False)
        mpa = mpa.fillna(0)
        mpa_slise.to_excel("mpa_1.xlsx", index=False)
        ### создаем список кодов KKS, типов прокладок, чтобы позже вписать в объединенные ячейки
        list_of_KKS = []
        for i in range(len(mpa_slise['Код KKS'])):
            list_of_KKS.append(mpa_slise['Код KKS'][i])
        list_of_laying_type = []
        for i in range(len(mpa_slise['Тип прокладки'])):
            list_of_laying_type.append(mpa_slise['Тип прокладки'][i])
        book = openpyxl.load_workbook("input/Templates/Template-MPA0001.xlsx")
        l = (mpa_slise.shape[1] - 2) * mpa_slise.shape[0]
        sheet = book['Sheet1']
        column_a = sheet['A']
        column_c = sheet['C']
        column_d = sheet['D']
        column_e = sheet['E']
        column_f = sheet['F']
        column_g = sheet['G']
        column_h = sheet['H']
        column_i = sheet['I']
        column_j = sheet['J']
        column_k = sheet['K']
        column_l = sheet['L']
        i = 3
        j = 0
        m = 0 # переменная для итерации по колонке количество из датафрейма pipeline_with_insulation
        while i <= l:
            column_g[i-1].value = mpa_slise['Код KKS'][j] + ' ' + mpa_slise['Тип прокладки'][j] + ' прокладка'
            column_g[i].value = (mpa_slise['mati_2m100_80mm * Количество'][j] * 1.236).round(3)
            column_d[i].value = 'Сборный'
            column_f[i].value = 'м' + chr(179)
            column_e[i].value = 'ГОСТ 21880-2011'
            column_c[i].value = 'Маты минераловатные прошивные марки 2М-100 в обкладке из стеклоткани марки Т-23 со всех сторон k=1.2,k=1.03 s=90 мм'
            column_l[i].value = 'Объем дан с учетом коэффициентов. Объем без учёта коэффициентов' + ' ' + str(
                (mpa_slise['mati_2m100_80mm * Количество'][j]).round(3)) +' м' + chr(179)
            column_g[i + 1].value = (mpa_slise['mati_2m100_70mm * Количество'][j] * 1.236).round(3)
            column_d[i + 1].value = 'Сборный'
            column_f[i + 1].value = 'м' + chr(179)
            column_e[i + 1].value = 'ГОСТ 21880-2011'
            column_c[i + 1].value = 'Маты минераловатные прошивные марки 2М-100 в обкладке из стеклоткани марки Т-23 со всех сторон k=1.2,k=1.03 s=80 мм'
            column_l[i + 1].value = 'Объем дан с учетом коэффициентов. Объем без учёта коэффициентов' + ' ' + str(
                (mpa_slise['mati_2m100_70mm * Количество'][j]).round(3)) + ' м' + chr(179)
            column_g[i + 2].value = (mpa_slise['mati_2m100_60mm * Количество'][j] * 1.236).round(3)
            column_d[i + 2].value = 'Сборный'
            column_f[i + 2].value = 'м' + chr(179)
            column_e[i + 2].value = 'ГОСТ 21880-2011'
            column_c[i + 2].value = 'Маты минераловатные прошивные марки 2М-100 в обкладке из стеклоткани марки Т-23 со всех сторон k=1.2,k=1.03 s=70 мм'
            column_l[i + 2].value = 'Объем дан с учетом коэффициентов. Объем без учёта коэффициентов' + ' ' + str(
                (mpa_slise['mati_2m100_60mm * Количество'][j]).round(3)) + ' м' + chr(179)
            column_g[i + 3].value = (mpa_slise['mati_2m100_50mm * Количество'][j] * 1.236).round(3)
            column_d[i + 3].value = 'Сборный'
            column_f[i + 3].value = 'м' + chr(179)
            column_e[i + 3].value = 'ГОСТ 21880-2011'
            column_c[i + 3].value = 'Маты минераловатные прошивные марки 2М-100 в обкладке из стеклоткани марки Т-23 со всех сторон k=1.2,k=1.03 s=60 мм'
            column_l[i + 3].value = 'Объем дан с учетом коэффициентов. Объем без учёта коэффициентов' + ' ' + str(
                (mpa_slise['mati_2m100_50mm * Количество'][j]).round(3)) + ' м' + chr(179)
            if armatura_presence == "да":
                column_g[i + 4].value = (mpa_slise['mati_2m100_40mm * Количество'][j] * 1.236).round(3)
                column_l[i + 4].value = 'Объем дан с учетом коэффициентов. Объем без учёта коэффициентов' + ' ' + str(
                    (mpa_slise['mati_2m100_40mm * Количество'][j]).round(3)) + ' м' + chr(179)
            else:
                column_g[i + 4].value = 0
                column_l[i + 4].value = 0
            column_d[i + 4].value = 'Сборный'
            column_f[i + 4].value = 'м' + chr(179)
            column_e[i + 4].value = 'ГОСТ 21880-2011'
            column_c[i + 4].value = 'Маты минераловатные прошивные марки 2М-100 в обкладке из стеклоткани марки Т-23 со всех сторон k=1.2,k=1.03 s=50 мм'
            column_g[i + 5].value = (mpa_slise['mati_m100_70mm * Количество'][j] * 1.236).round(3)
            column_d[i + 5].value = 'Сборный'
            column_f[i + 5].value = 'м' + chr(179)
            column_e[i + 5].value = 'ГОСТ 21880-2011'
            column_c[i + 5].value = "Маты минераловатные прошивные марки М-100 без обкладок k=1.2,k=1.03 s=80 мм"
            column_l[i + 5].value = 'Объем дан с учетом коэффициентов. Объем без учёта коэффициентов' + ' ' + str(
                (mpa_slise['mati_m100_70mm * Количество'][j]).round(3)) + ' м' + chr(179)
            column_g[i + 6].value = (mpa_slise['mati_m100_60mm * Количество'][j] * 1.236).round(3)
            column_d[i + 6].value = 'Сборный'
            column_f[i + 6].value = 'м' + chr(179)
            column_e[i + 6].value = 'ГОСТ 21880-2011'
            column_c[i + 6].value = "Маты минераловатные прошивные марки М-100 без обкладок k=1.2,k=1.03 s=70 мм"
            column_l[i + 6].value = 'Объем дан с учетом коэффициентов. Объем без учёта коэффициентов' + ' ' + str(
                (mpa_slise['mati_m100_60mm * Количество'][j]).round(3)) + ' м' + chr(179)
            column_g[i + 7].value = (mpa_slise['mati_m100_50mm * Количество'][j] * 1.236).round(3)
            column_d[i + 7].value = 'Сборный'
            column_f[i + 7].value = 'м' + chr(179)
            column_e[i + 7].value = 'ГОСТ 21880-2011'
            column_c[i + 7].value = "Маты минераловатные прошивные марки М-100 без обкладок k=1.2,k=1.03 s=60 мм"
            column_l[i + 7].value = 'Объем дан с учетом коэффициентов. Объем без учёта коэффициентов' + ' ' + str(
                (mpa_slise['mati_m100_50mm * Количество'][j]).round(3)) + ' м' + chr(179)
            column_c[i + 8].value = 'Cтеклоткань марки Т-23'
            column_d[i + 8].value = 'Т-23'
            if armatura_presence == "да":
                column_g[i + 8].value = ((mpa_slise['T_23_2m100_80mm'][j] + mpa_slise['T_23_2m100_70mm'][j] +
                mpa_slise['T_23_2m100_60mm'][j] + mpa_slise['T_23_2m100_50mm'][j] + mpa_slise['T_23_2m100_40mm'][j])
                                     * 1.236).round(3)
            else:
                column_g[i + 8].value = ((mpa_slise['T_23_2m100_80mm'][j] + mpa_slise['T_23_2m100_70mm'][j] +
                                          mpa_slise['T_23_2m100_60mm'][j] + mpa_slise['T_23_2m100_50mm'][j]) * 1.236).round(3)
            column_g[i + 8].number_format = "0.000"
            column_l[i + 8].value = '(для матов в обкладках)'
            column_f[i + 8].value = 'м' + chr(178)
            column_e[i + 8].value = 'по типу ТУ 6-11-231-76'
            column_c[i + 9].value = 'Cтеклоткань марки Т-23'
            column_d[i + 9].value = 'Т-23'
            column_g[i + 9].value = ((mpa_slise['T_23_m100_70mm'][j] + mpa_slise['T_23_m100_60mm'][j] +
                                     mpa_slise['T_23_m100_50mm'][j]) * 1.16 * 1.236).round(3)
            column_g[i + 9].number_format = "0.000"
            column_l[i + 9].value = '(для матов без обкладок)'
            column_f[i + 9].value = 'м' + chr(178)
            column_e[i + 9].value = 'по типу ТУ 6-11-231-76'
            column_c[i + 10].value = 'Покрытие защитное из стеклоткани марки Т-23'
            column_g[i + 10].value = (mpa_slise['protective_cover_T_23'][j] * pipeline_with_insulation['Количество'][m] * 1.16).round(3)
            column_g[i + 10].number_format = "0.000"
            column_d[i + 10].value = 'Т-23'
            column_f[i + 10].value = 'м' + chr(178)
            column_e[i + 10].value = 'по типу ТУ 6-11-231-76'
            column_c[i + 11].value = 'Cталь тонколистовая оцинкованная s=0.8 мм'
            column_g[i + 11].value = (mpa_slise['steel_0_8mm * Количество'][j] * 1.22).round(3)
            column_g[i + 11].number_format = "0.000"
            column_l[i + 11].value = "Объем дан с учетом коэффициентов. Объем без учёта коэффициентов" + " " +\
                                     str(mpa_slise['steel_0_8mm * Количество'][j].round(3)) + ' м' + chr(178) +' в констр.'
            column_d[i + 11].value = 'Оцинкованная сталь'
            column_f[i + 11].value = 'м' + chr(178)
            column_e[i + 11].value = 'ГОСТ 14918-80'
            column_c[i + 12].value = 'Cталь тонколистовая оцинкованная s=0.5 мм'
            column_g[i + 12].value = (mpa_slise['steel_0_5mm * Количество'][j] * 1.22).round(3)
            column_g[i + 12].number_format = "0.000"
            column_l[i + 12].value = "Объем дан с учетом коэффициентов. Объем без учёта коэффициентов" + " " + \
                                     str(mpa_slise['steel_0_5mm * Количество'][j].round(3)) + ' м' + chr(178) +' в констр.'
            column_d[i + 12].value = 'Оцинкованная сталь'
            column_f[i + 12].value = 'м' + chr(178)
            column_e[i + 12].value = 'ГОСТ 14918-80'
            column_c[i + 13].value = 'Скоба опорная'
            column_g[i + 13].value = mpa_slise['Скоба опорная * Количество'][j].round(0)
            column_g[i + 13].number_format = "0"
            column_e[i + 13].value = 'ГОСТ 14918-80'
            column_d[i + 13].value = 'Оцинкованная сталь'
            column_f[i + 13].value = 'шт.'
            column_c[i + 14].value = 'Пряжка тип I-О'
            column_g[i + 14].value = mpa_slise['Пряжка тип I-О * Количество'][j].round(0)
            column_g[i + 14].number_format = "0"
            column_d[i + 14].value = 'Сборный'
            column_f[i + 14].value = 'шт.'
            column_e[i + 14].value = 'по типу ТУ 36.16.22-64-92'
            column_c[i + 15].value = 'Лента стальная упаковочная 0,7х20 мм'
            column_g[i + 15].value = mpa_slise['Лента 0,7х20 * Количество'][j].round(3)
            column_f[i + 15].value = 'кг'
            column_d[i + 15].value = 'Углеродистая сталь'
            column_e[i + 15].value = 'ГОСТ 3560-73'
            column_c[i + 16].value = 'Проволока 2-0-Ч'
            column_g[i + 16].value = mpa_slise['Проволока 2-0-Ч * Количество'][j].round(3)
            column_d[i + 16].value = 'Углеродистая сталь'
            column_f[i + 16].value = 'кг'
            column_e[i + 16].value = 'ГОСТ 3282-74'
            column_c[i + 17].value = 'Винт самонарезающий'
            column_g[i + 17].value = mpa_slise['Винт самонарезной * Количество'][j].round(0)
            column_g[i + 17].number_format = "0"
            column_d[i + 17].value = 'Сборный'
            column_f[i + 17].value = 'шт.'
            column_e[i + 17].value = 'ГОСТ 10621-80'
            column_c[i + 18].value = 'Проволока 0,8-0-Ч'
            column_g[i + 18].value = mpa_slise['Проволока 0,8-0-Ч * Количество'][j].round(3)
            column_d[i + 18].value = 'Углеродистая сталь'
            column_f[i + 18].value = 'кг'
            column_e[i + 18].value = 'ГОСТ 3282-74'
            column_c[i + 19].value = 'Лак БТ-577'
            column_g[i + 19].value = mpa_slise['Лак БТ-7,кг * Количество'][j].round(3)
            column_l[i + 19].value = str(mpa_slise['Лак БТ-7,м2 * Количество'][j]) + ' м' + chr(178)
            column_d[i + 19].value = 'Сборный'
            column_f[i + 19].value = 'кг'
            column_e[i + 19].value = 'ГОСТ 5631-79'
            column_c[i + 20].value = 'Лента АД1.08х20'
            if armatura_presence == "да":
                column_g[i + 20].value = mpa_slise['Лента АД1.08х20 * Количество'][j].round(3)
                column_g[i + 21].value = mpa_slise['Пряжка тип II-А * Количество'][j].round(0)
                column_g[i + 22].value = mpa_slise['Заклепка комбинированная ЗК-12-4,5 * Количество'][j].round(0)
                column_g[i + 23].value = mpa_slise['Крючок * Количество'][j].round(3)
                column_g[i + 24].value = mpa_slise['Серьга * Количество'][j].round(3)
                column_g[i + 25].value = mpa_slise['Рычаг * Количество'][j].round(3)
                column_g[i + 26].value = mpa_slise['Основание * Количество'][j].round(3)
                column_g[i + 27].value = mpa_slise['Заклепка 4x24.37 * Количество'][j].round(3)
            else:
                column_g[i + 20].value = 0
                column_g[i + 21].value = 0
                column_g[i + 22].value = 0
                column_g[i + 23].value = 0
                column_g[i + 24].value = 0
                column_g[i + 25].value = 0
                column_g[i + 26].value = 0
                column_g[i + 27].value = 0
            column_d[i + 20].value = 'АД1.08х20'
            column_f[i + 20].value = 'кг'
            column_e[i + 20].value = 'ГОСТ 13726-97'
            column_c[i + 21].value = 'Пряжка тип II-А'
            column_g[i + 21].number_format = "0"
            column_d[i + 21].value = 'Сборный'
            column_f[i + 21].value = 'шт.'
            column_e[i + 21].value = 'по типу ТУ 36.16.22-64-92'
            column_c[i + 22].value = 'Заклепка комбинированная ЗК-12-4,5'
            column_g[i + 22].number_format = "0"
            column_d[i + 22].value = 'Сборный'
            column_f[i + 22].value = 'шт.'
            column_e[i + 22].value = 'по типу ТУ 36-2088-85'
            column_c[i + 23].value = 'Крючок'
            column_d[i + 23].value = 'Углеродистая сталь'
            column_f[i + 23].value = 'кг'
            column_e[i + 23].value = 'ГОСТ 19904-90'
            column_l[i + 23].value = 'Для крепления арматуры DN>200'
            column_c[i + 24].value = 'Серьга'
            column_d[i + 24].value = 'Углеродистая сталь'
            column_f[i + 24].value = 'кг'
            column_e[i + 24].value = 'ГОСТ 3282-74'
            column_l[i + 24].value = 'Для крепления арматуры DN>200'
            column_c[i + 25].value = 'Рычаг'
            column_d[i + 25].value = 'Углеродистая сталь'
            column_f[i + 25].value = 'кг'
            column_e[i + 25].value = 'ГОСТ 19904-90'
            column_l[i + 25].value = 'Для крепления арматуры DN>200'
            column_c[i + 26].value = 'Основание'
            column_d[i + 26].value = 'Углеродистая сталь'
            column_f[i + 26].value = 'кг'
            column_e[i + 26].value = 'ГОСТ 19904-90'
            column_l[i + 26].value = 'Для крепления арматуры DN>200'
            column_c[i + 27].value = 'Заклепка 4x24.37'
            column_d[i + 27].value = 'Сборный'
            column_f[i + 27].value = 'кг'
            column_e[i + 27].value = 'ГОСТ 10299-80'
            column_l[i + 27].value = 'Для крепления арматуры DN>200'
            column_c[i + 28].value = 'Лента 2х30 Ст3пс'
            column_g[i + 28].value = mpa_slise['Лента 2х30 * Длина вертикальных участков'][j].round(3)
            column_d[i + 28].value = 'Углеродистая сталь'
            column_f[i + 28].value = 'кг'
            column_e[i + 28].value = 'ГОСТ 6009-74'
            column_l[i + 28].value = str((mpa_slise['mati_2m100_80mm * Длина вертикальных участков'][j] +
                                      mpa_slise['mati_2m100_70mm * Длина вертикальных участков'][j] +
                                      mpa_slise['mati_2m100_60mm * Длина вертикальных участков'][j] +
                                      mpa_slise['mati_2m100_50mm * Длина вертикальных участков'][j] +
                                      mpa_slise['mati_m100_70mm * Длина вертикальных участков'][j] +
                                      mpa_slise['mati_m100_60mm * Длина вертикальных участков'][j] +
                                      mpa_slise['mati_m100_50mm * Длина вертикальных участков'][j]).round(3)) + ' м' + chr(179)
            column_c[i + 29].value = 'Лента 3х30 Ст3пс'
            column_g[i + 29].value = mpa_slise['Лента 3х30 * Длина вертикальных участков'][j].round(3)
            column_d[i + 29].value = 'Углеродистая сталь'
            column_f[i + 29].value = 'кг'
            column_e[i + 29].value = 'ГОСТ 6009-74'
            column_c[i + 30].value = 'Уголок 30х30x3'
            column_g[i + 30].value = mpa_slise['Уголок 30х30x3 * Длина вертикальных участков'][j].round(3)
            column_d[i + 30].value = 'Углеродистая сталь'
            column_f[i + 30].value = 'кг'
            column_e[i + 30].value = 'ГОСТ 8509-93'
            column_c[i + 31].value = 'Скоба навесная'
            column_g[i + 31].value = mpa_slise['Скоба навесная * Длина вертикальных участков'][j].round(0)
            column_d[i + 31].value = 'Оцинкованная сталь'
            column_f[i + 31].value = 'шт.'
            column_e[i + 31].value = 'ГОСТ 14918-80'
            column_c[i + 32].value = 'Лист АД1.Н-0.8'
            column_g[i + 32].value = mpa_slise['Лист АД1.Н-0.8 * Длина вертикальных участков'][j].round(3)
            column_d[i + 32].value = 'Алюминий'
            column_f[i + 32].value = 'кг'
            column_e[i + 32].value = 'ГОСТ 21631-76'
            column_c[i + 33].value = 'Болт М8x30.36.019'
            column_g[i + 33].value = mpa_slise['Болт М8 * Длина вертикальных участков'][j].round(3)
            column_d[i + 33].value = 'Углеродистая сталь'
            column_f[i + 33].value = 'кг'
            column_e[i + 33].value = 'ГОСТ 7798-70'
            column_c[i + 34].value = 'Болт М12x50.36.019'
            column_g[i + 34].value = mpa_slise['Болт М12 * Длина вертикальных участков'][j].round(3)
            column_d[i + 34].value = 'Углеродистая сталь'
            column_f[i + 34].value = 'кг'
            column_e[i + 34].value = 'ГОСТ 7798-70'
            column_c[i + 35].value = 'Гайка М8.4.019'
            column_g[i + 35].value = mpa_slise['Гайка М8 * Длина вертикальных участков'][j].round(3)
            column_d[i + 35].value = 'Углеродистая сталь'
            column_f[i + 35].value = 'кг'
            column_e[i + 35].value = 'ГОСТ 5915-70'
            column_c[i + 36].value = 'Гайка М12.4.019'
            column_g[i + 36].value = mpa_slise['Гайка М12 * Длина вертикальных участков'][j].round(3)
            column_d[i + 36].value = 'Углеродистая сталь'
            column_f[i + 36].value = 'кг'
            column_e[i + 36].value = 'ГОСТ 5915-70'

            i += (mpa_slise.shape[1] - 2)
            j += 1
            m += 1
            a = i
        ### Убираем пустые ячейки

        list_of_merge_rows = []
        n = 1
        for cell in sheet['G'][2:a + 1]:
            if cell.value == 0 or cell.value == None:
                sheet.delete_rows(cell.row)
            if type(cell.value) == str:
                list_of_merge_rows.append(cell.row)
                sheet.merge_cells(start_row=cell.row, start_column=1, end_row=cell.row, end_column=12)
        j = 0
        for i in range(len(list_of_merge_rows)):
            sheet['A'][list_of_merge_rows[i] - 1].value = list_of_KKS[j] + " " +list_of_laying_type[j] + " прокладка"
            sheet['A'][list_of_merge_rows[i] - 1].font = Font(name='Arial', size=14, bold=True)
            right = Side(border_style="medium", color='000000')
            border = Border(right = right)
            sheet['L'][list_of_merge_rows[i] -1].border = border
            sheet.row_dimensions[list_of_merge_rows[i]].height = 23
            j += 1
        # print(sheet.max_row)
        n = 1
        for cell in sheet['A'][3:a + 1]:
            if cell.value == None:
                cell.value = n
                n += 1

        sheet.oddFooter.left.text = КодKKS + "-MPA0001" + '\n' + Архивный
        return book

    book_mpa = mpa_write()
    book_mpa.save(f"output/{КодKKS}-MPA0001.xlsx")
## Список тегов для заполения шаблона:пункт графика, Код КодKKSдокумента РД, Код КодKKSдокумента спецификации РД,
## Код КодKKSдокумента спецификации арматуры РД, Код КодKKSдокумента документа, Архивный номер, номер МДС, архивный номер см
    def mdb_write():
        mdb = DocxTemplate("input/Templates/Template-MDB0001.docx")
        context = {'КодKKSРД': КодKKSдокумента, 'КодKKSспецификацииРД': КодKKSспецификацииРД,
                   'КодKKSспецификацииарматурыРД': КодKKSспецификацииарматурыРД,
                   'КодKKSдокумента': КодKKS,
                   'номерМДС': номерМДС, 'Архивный': Архивный, 'пункт_графика': пункт_графика}
        mdb.render(context)
        return mdb

    mdb = mdb_write()
    mdb.save(f"output/{КодKKS}-MDB0001.docx")

    def maa_write():
        maa = DocxTemplate("input/Templates/Template-MAA0001.docx")
        context = {'КодKKSдокумента' : КодKKS, 'Наименование_работы' : Наименование_работы,
        'Исполнитель_должн' : Исполнитель_должн, 'Исполнитель' : Исполнитель, 'Архивный' : Архивный}
        maa.render(context)
        return maa

    maa = maa_write()
    maa.save(f"output/{КодKKS}-MAA0001.docx")

    def maz_write():
        maz = DocxTemplate("input/Templates/Template-MAZ0001.docx")
        context = {'КодKKSдокумента' : КодKKS , 'Архивный' : Архивный}
        maz.render(context)
        return maz

    maz = maz_write()
    maz.save(f"output/{КодKKS}-MAZ0001.docx")

    def mab_write():
        mab = DocxTemplate("input/Templates/Template-MAB0001.docx")
        context = {'КодKKSдокумента' : КодKKS, 'Архивный' : Архивный}
        mab.render(context)
        return mab

    mab = mab_write()
    mab.save(f"output/{КодKKS}-MAB0001.docx")


    def ls_write():
        ls = DocxTemplate("input/Templates/Template-LS.docx")
        context = {'КодKKSдокумента' : КодKKS, 'Исполнитель' : Исполнитель, 'Архивный' : Архивный}
        ls.render(context)
        return ls

    ls = ls_write()
    ls.save("output/Лист согласования.docx")

    def checklist_write():
        checklist = DocxTemplate("input/Templates/Template-checklist.docx")
        context = {'КодKKSдокумента' : КодKKSдокумента, 'Исполнитель' : Исполнитель, 'Архивный' : Архивный}
        checklist.render(context)
        return checklist

    checklist = checklist_write()
    checklist.save("output/Чек лист.docx")

    def estimate_task_write():
        estimate_task = DocxTemplate("input/Templates/Template-estimate_task.docx")
        context = {'КодKKSдокумента' : КодKKS, 'пункт_графика' : пункт_графика}
        estimate_task.render(context)
        return estimate_task

    estimate_task = estimate_task_write()
    estimate_task.save("output/Задание на сметы.docx")

    def ul_write():
        ul = DocxTemplate("input/Templates/Template-ul.docx")
        context = {'дата' : дата, 'Архивный' : Архивный, 'Исполнитель' : Исполнитель, 'Наименование_работы' : Наименование_работы}
        ul.render(context)
        return ul

    ul = ul_write()
    ul.save("output/Удостоверяющий лист.docx")


    def protocol_write():
        protocol = DocxTemplate("input/Templates/Template-protocol.docx")
        context = {'дата' : дата, 'Архивный' : Архивный, 'Исполнитель' : Исполнитель, 'Наименование_работы' : Наименование_работы,
    'Исполнитель_должн' : Исполнитель_должн, 'КодKKSдокумента' : КодKKS}
        protocol.render(context)
        return protocol

    protocol = protocol_write()
    protocol.save("output/Протокол.docx")

#output()

if __name__ == '__main__':
    start_time = time.time()
    pipe_parsed = pipelines_parse()
    armatura_presence = armatura_presence()
    armatura_parsed = parse_armatura()
    armatura_pos1_5 = armatura_pos1_5()
    armatura_fasteners = armatura_fasteners()
    pipeline_fasteners = pipeline_fasteners()
    ascent_list = ascents()
    pipe_parsed = pipeline_insulation_thickness()
    pipeline_with_insulation = pipeline_insun_preporation()
    valves_insulation = valves_insulation()
    mpd = mpd_preporation()
    list_of_insulation = insulation_list_prep()
    output()
    print("--- %s seconds ---" % (time.time() - start_time))



























